import csv
import io
import secrets
from datetime import datetime
from functools import wraps
import os

from flasgger import Swagger

from flask import (
    Flask,
    Response,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)

from database import (
    PRIORIDADE_CRITICA,
    PRIORIDADE_ORDEM_SQL,
    PRIORIDADE_SLUGS,
    PRIORIDADES,
    STATUS_ABERTA,
    STATUS_CANCELADA,
    STATUS_CONCLUIDA,
    STATUS_EM_ANDAMENTO,
    TODOS_STATUS,
    authenticate_user,
    get_all_users_with_stats,
    get_db_connection,
    initialize_database,
)


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "sgdi-dev-key-troque-em-producao-2024")

Swagger(app, template={
    "swagger": "2.0",
    "info": {
        "title": "SGDI API",
        "description": (
            "API REST para integração com o Sistema de Gestão de Demandas Internas.\n\n"
            "**Autenticação:** inclua o header `X-API-Key` em todas as requisições.\n"
            "Gere sua chave em `/api/keys` (login necessário)."
        ),
        "version": "1.0.0",
    },
    "securityDefinitions": {
        "ApiKeyAuth": {"type": "apiKey", "in": "header", "name": "X-API-Key"}
    },
    "security": [{"ApiKeyAuth": []}],
    "tags": [
        {"name": "Demandas", "description": "Consulta e manipulação de demandas"},
        {"name": "Usuários",  "description": "Listagem de usuários ativos"},
        {"name": "Comentários", "description": "Comentários em demandas"},
    ],
}, config={
    "headers": [],
    "specs": [{"endpoint": "apispec", "route": "/apispec.json",
               "rule_filter": lambda r: r.rule.startswith("/api/v1"),
               "model_filter": lambda t: True}],
    "static_url_path": "/flasgger_static",
    "swagger_ui": True,
    "specs_route": "/apidocs",
})

ORDENACOES = {
    "prioridade": f"{PRIORIDADE_ORDEM_SQL}, d.data_criacao DESC",
    "recentes": "d.data_criacao DESC",
}
ORDENACAO_LABELS = {
    "prioridade": "Prioridade primeiro",
    "recentes": "Mais recentes primeiro",
}
DIAS_ALERTA_PARADA = 7

initialize_database()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario_id" not in session:
            flash("Faça login para continuar.")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def api_key_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        chave = request.headers.get("X-API-Key", "").strip()
        if not chave:
            return jsonify({"success": False, "error": "Header X-API-Key ausente"}), 401
        conn = get_db()
        try:
            row = conn.execute(
                "SELECT id FROM api_keys WHERE chave = ? AND ativo = 1", (chave,)
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return jsonify({"success": False, "error": "Chave de API inválida ou desativada"}), 403
        return f(*args, **kwargs)
    return decorated


def _api_ok(data, total=None):
    payload = {"success": True, "data": data}
    if total is not None:
        payload["meta"] = {"total": total}
    return jsonify(payload)


def _api_err(msg, status):
    return jsonify({"success": False, "error": msg}), status


@app.context_processor
def inject_globals():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return {
        "csrf_token": session["csrf_token"],
        "usuario_logado": {
            "id": session.get("usuario_id"),
            "nome": session.get("usuario_nome"),
            "username": session.get("usuario_username"),
        } if "usuario_id" in session else None,
    }


def _validate_csrf():
    token = request.form.get("csrf_token", "")
    if not token or token != session.get("csrf_token"):
        abort(403)


def get_db():
    return get_db_connection()


@app.template_filter("priority_slug")
def priority_slug(value):
    return PRIORIDADE_SLUGS.get(value, "neutra")


@app.template_filter("display_datetime")
def display_datetime(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y às %H:%M")
    except (TypeError, ValueError):
        return value or ""


def normalize_priority(prioridade):
    return prioridade if prioridade in PRIORIDADES else ""


def normalize_order(ordenacao):
    return ordenacao if ordenacao in ORDENACOES else "prioridade"


def _escape_like(term):
    return term.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def build_demand_context(
    status,
    filtro_prioridade="",
    filtro_usuario_id=None,
    termo_busca="",
    ordenacao="prioridade",
):
    filtro_normalizado = normalize_priority(filtro_prioridade)
    ordenacao_normalizada = normalize_order(ordenacao)

    conn = get_db()
    try:
        clauses = ["d.status = ?"]
        params = [status]

        if filtro_normalizado:
            clauses.append("d.prioridade = ?")
            params.append(filtro_normalizado)

        if filtro_usuario_id:
            try:
                clauses.append("d.responsavel_id = ?")
                params.append(int(filtro_usuario_id))
            except (ValueError, TypeError):
                pass

        if termo_busca:
            termo_like = f"%{_escape_like(termo_busca)}%"
            clauses.append(
                "(d.titulo LIKE ? ESCAPE '\\' "
                "OR d.descricao LIKE ? ESCAPE '\\' "
                "OR d.solicitante LIKE ? ESCAPE '\\')"
            )
            params.extend([termo_like, termo_like, termo_like])

        query = (
            "SELECT d.*, u.username AS usuario_username, "
            "COALESCE(resp.nome, '') AS responsavel_nome "
            "FROM demandas d "
            "JOIN usuarios u ON u.id = d.usuario_id "
            "LEFT JOIN usuarios resp ON resp.id = d.responsavel_id "
            f"WHERE {' AND '.join(clauses)} "
            f"ORDER BY {ORDENACOES[ordenacao_normalizada]}"
        )
        demandas = conn.execute(query, tuple(params)).fetchall()
        usuarios = conn.execute(
            "SELECT id, nome FROM usuarios ORDER BY nome"
        ).fetchall()
    finally:
        conn.close()

    demandas_enriquecidas = enrich_demands(demandas)

    return {
        "demandas": demandas_enriquecidas,
        "filtro_atual": filtro_normalizado,
        "filtro_usuario_id": filtro_usuario_id or "",
        "ordenacao_atual": ordenacao_normalizada,
        "resumo": build_summary(
            demandas_enriquecidas,
            filtro_normalizado,
            ordenacao_normalizada,
        ),
        "prioridades": PRIORIDADES,
        "termo_busca": termo_busca,
        "usuarios": usuarios,
    }


def enrich_demands(demandas):
    agora = datetime.now()
    resultado = []
    for demanda in demandas:
        d = dict(demanda)
        try:
            data_criacao = datetime.strptime(d["data_criacao"], "%Y-%m-%d %H:%M:%S")
            dias_parada = (agora - data_criacao).days
        except (ValueError, TypeError):
            dias_parada = 0
        d["dias_parada"] = dias_parada
        d["alerta_parada"] = d.get("status") == STATUS_ABERTA and dias_parada >= DIAS_ALERTA_PARADA
        resultado.append(d)
    return resultado


def _registrar_historico(conn, demanda_id, status_anterior, status_novo, autor, now):
    conn.execute(
        "INSERT INTO historico_status (demanda_id, status_anterior, status_novo, autor, data) "
        "VALUES (?, ?, ?, ?, ?)",
        (demanda_id, status_anterior, status_novo, autor, now),
    )


def build_summary(demandas, filtro_prioridade, ordenacao):
    contagem = {p: 0 for p in PRIORIDADES}
    for d in demandas:
        if d.get("prioridade") in contagem:
            contagem[d["prioridade"]] += 1
    return {
        "total": len(demandas),
        "alta": contagem[PRIORIDADES[0]],
        "media": contagem[PRIORIDADES[1]],
        "baixa": contagem[PRIORIDADES[2]],
        "alertas": sum(1 for d in demandas if d.get("alerta_parada")),
        "filtro_label": filtro_prioridade or "Todas as prioridades",
        "ordenacao_label": ORDENACAO_LABELS[ordenacao],
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if "usuario_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        senha = request.form.get("senha", "").strip()

        usuario = authenticate_user(username, senha)
        if usuario:
            session.clear()
            session["usuario_id"] = usuario["id"]
            session["usuario_nome"] = usuario["nome"]
            session["usuario_username"] = usuario["username"]
            session["csrf_token"] = secrets.token_hex(32)
            flash(f"Bem-vindo, {usuario['nome']}!")
            return redirect(url_for("dashboard"))

        flash("Usuário ou senha incorretos.")

    return render_template("login.html")


@app.route("/logout", methods=["POST"])
@login_required
def logout():
    _validate_csrf()
    session.clear()
    flash("Sessão encerrada com sucesso.")
    return redirect(url_for("login"))


@app.route("/demandas")
@login_required
def index():
    contexto = build_demand_context(
        STATUS_ABERTA,
        filtro_prioridade=request.args.get("prioridade", "").strip(),
        filtro_usuario_id=request.args.get("usuario_id", "").strip() or None,
        ordenacao=request.args.get("ordenacao", "prioridade"),
    )
    contexto["rota_listagem"] = "index"
    return render_template("index.html", **contexto)


@app.route("/concluidas")
@login_required
def concluidas():
    contexto = build_demand_context(
        STATUS_CONCLUIDA,
        filtro_prioridade=request.args.get("prioridade", "").strip(),
        filtro_usuario_id=request.args.get("usuario_id", "").strip() or None,
        ordenacao=request.args.get("ordenacao", "prioridade"),
    )
    return render_template("concluidas.html", **contexto)


@app.route("/nova_demanda", methods=["GET", "POST"])
@login_required
def nova_demanda():
    conn = get_db()
    try:
        usuarios = conn.execute("SELECT id, nome FROM usuarios ORDER BY nome").fetchall()
    finally:
        conn.close()

    form_data = {"titulo": "", "descricao": "", "prioridade": "", "data_prevista": "", "responsavel_id": ""}

    if request.method == "POST":
        _validate_csrf()
        data_prevista_raw = request.form.get("data_prevista", "").strip()
        data_prevista = None
        if data_prevista_raw:
            try:
                datetime.strptime(data_prevista_raw, "%Y-%m-%d")
                data_prevista = data_prevista_raw
            except ValueError:
                pass
        responsavel_id_raw = request.form.get("responsavel_id", "").strip()
        responsavel_id = None
        if responsavel_id_raw:
            try:
                responsavel_id = int(responsavel_id_raw)
            except ValueError:
                pass
        form_data = {
            "titulo": request.form.get("titulo", "").strip(),
            "descricao": request.form.get("descricao", "").strip(),
            "prioridade": request.form.get("prioridade", "").strip(),
            "data_prevista": data_prevista or "",
            "responsavel_id": responsavel_id_raw,
        }

        if not form_data["titulo"] or not form_data["descricao"]:
            flash("Todos os campos são obrigatórios.")
            return render_template("nova_demanda.html", prioridades=PRIORIDADES, form_data=form_data, usuarios=usuarios)

        if form_data["prioridade"] not in PRIORIDADES:
            flash("Selecione uma prioridade válida.")
            return render_template("nova_demanda.html", prioridades=PRIORIDADES, form_data=form_data, usuarios=usuarios)

        conn = get_db()
        try:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor = conn.execute(
                """
                INSERT INTO demandas (
                    titulo, descricao, solicitante, usuario_id,
                    prioridade, status, data_criacao, data_prevista, responsavel_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    form_data["titulo"],
                    form_data["descricao"],
                    session["usuario_nome"],
                    session["usuario_id"],
                    form_data["prioridade"],
                    STATUS_ABERTA,
                    now,
                    data_prevista,
                    responsavel_id,
                ),
            )
            _registrar_historico(conn, cursor.lastrowid, None, STATUS_ABERTA, session["usuario_nome"], now)
            conn.commit()
        finally:
            conn.close()

        flash("Demanda criada com sucesso.")
        return redirect(url_for("index"))

    return render_template("nova_demanda.html", prioridades=PRIORIDADES, form_data=form_data, usuarios=usuarios)


@app.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar(id):
    conn = get_db()
    try:
        usuarios = conn.execute("SELECT id, nome FROM usuarios ORDER BY nome").fetchall()
        demanda = conn.execute(
            "SELECT d.*, u.username AS usuario_username, "
            "COALESCE(resp.nome, '') AS responsavel_nome "
            "FROM demandas d "
            "JOIN usuarios u ON u.id = d.usuario_id "
            "LEFT JOIN usuarios resp ON resp.id = d.responsavel_id "
            "WHERE d.id = ?",
            (id,),
        ).fetchone()
        if not demanda:
            abort(404)

        if session["usuario_id"] != demanda["usuario_id"]:
            flash("Apenas o solicitante da demanda pode editá-la.")
            return redirect(url_for("detalhes", id=id))

        if request.method == "POST":
            _validate_csrf()
            titulo = request.form.get("titulo", "").strip()
            descricao = request.form.get("descricao", "").strip()
            prioridade = request.form.get("prioridade", "").strip()
            data_prevista_raw = request.form.get("data_prevista", "").strip()
            data_prevista = None
            if data_prevista_raw:
                try:
                    datetime.strptime(data_prevista_raw, "%Y-%m-%d")
                    data_prevista = data_prevista_raw
                except ValueError:
                    pass
            responsavel_id_raw = request.form.get("responsavel_id", "").strip()
            responsavel_id = None
            if responsavel_id_raw:
                try:
                    responsavel_id = int(responsavel_id_raw)
                except ValueError:
                    pass

            demanda_atualizada = dict(demanda)
            demanda_atualizada.update({
                "titulo": titulo,
                "descricao": descricao,
                "prioridade": prioridade or demanda["prioridade"],
                "data_prevista": data_prevista_raw or demanda.get("data_prevista"),
                "responsavel_id": responsavel_id if responsavel_id is not None else demanda.get("responsavel_id"),
            })

            if not titulo or not descricao:
                flash("Título e descrição são obrigatórios.")
                return render_template("editar.html", demanda=demanda_atualizada, prioridades=PRIORIDADES, usuarios=usuarios)

            if prioridade not in PRIORIDADES:
                flash("Selecione uma prioridade válida.")
                return render_template("editar.html", demanda=demanda_atualizada, prioridades=PRIORIDADES, usuarios=usuarios)

            conn.execute(
                "UPDATE demandas SET titulo = ?, descricao = ?, prioridade = ?, "
                "data_prevista = ?, responsavel_id = ? WHERE id = ?",
                (titulo, descricao, prioridade, data_prevista, responsavel_id, id),
            )
            conn.commit()
            flash("Demanda atualizada.")
            return redirect(url_for("detalhes", id=id))

        return render_template("editar.html", demanda=demanda, prioridades=PRIORIDADES, usuarios=usuarios)
    finally:
        conn.close()


@app.route("/concluir/<int:id>", methods=["POST"])
@login_required
def concluir(id):
    _validate_csrf()
    conn = get_db()
    try:
        demanda = conn.execute("SELECT * FROM demandas WHERE id = ?", (id,)).fetchone()
        if not demanda:
            abort(404)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            "UPDATE demandas SET status = ?, data_conclusao = ? WHERE id = ?",
            (STATUS_CONCLUIDA, now, id),
        )
        _registrar_historico(conn, id, demanda["status"], STATUS_CONCLUIDA, session["usuario_nome"], now)
        conn.commit()
    finally:
        conn.close()
    flash("Demanda marcada como concluída.")
    return redirect(url_for("index"))


@app.route("/reabrir/<int:id>", methods=["POST"])
@login_required
def reabrir(id):
    _validate_csrf()
    conn = get_db()
    try:
        demanda = conn.execute("SELECT * FROM demandas WHERE id = ?", (id,)).fetchone()
        if not demanda:
            abort(404)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("UPDATE demandas SET status = ? WHERE id = ?", (STATUS_ABERTA, id))
        _registrar_historico(conn, id, demanda["status"], STATUS_ABERTA, session["usuario_nome"], now)
        conn.commit()
    finally:
        conn.close()
    flash("Demanda reaberta.")
    return redirect(url_for("index"))


@app.route("/andamento/<int:id>", methods=["POST"])
@login_required
def andamento(id):
    _validate_csrf()
    conn = get_db()
    try:
        demanda = conn.execute("SELECT * FROM demandas WHERE id = ?", (id,)).fetchone()
        if not demanda:
            abort(404)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("UPDATE demandas SET status = ? WHERE id = ?", (STATUS_EM_ANDAMENTO, id))
        _registrar_historico(conn, id, demanda["status"], STATUS_EM_ANDAMENTO, session["usuario_nome"], now)
        conn.commit()
    finally:
        conn.close()
    flash("Demanda em andamento.")
    return redirect(url_for("detalhes", id=id))


@app.route("/cancelar/<int:id>", methods=["POST"])
@login_required
def cancelar(id):
    _validate_csrf()
    conn = get_db()
    try:
        demanda = conn.execute("SELECT * FROM demandas WHERE id = ?", (id,)).fetchone()
        if not demanda:
            abort(404)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("UPDATE demandas SET status = ? WHERE id = ?", (STATUS_CANCELADA, id))
        _registrar_historico(conn, id, demanda["status"], STATUS_CANCELADA, session["usuario_nome"], now)
        conn.commit()
    finally:
        conn.close()
    flash("Demanda cancelada.")
    return redirect(url_for("detalhes", id=id))


@app.route("/deletar/<int:id>", methods=["POST"])
@login_required
def deletar(id):
    _validate_csrf()
    conn = get_db()
    try:
        demanda = conn.execute("SELECT * FROM demandas WHERE id = ?", (id,)).fetchone()
        if not demanda:
            abort(404)
        destino = "concluidas" if demanda["status"] == STATUS_CONCLUIDA else "index"
        if session["usuario_id"] != demanda["usuario_id"]:
            flash("Apenas o solicitante da demanda pode deletá-la.")
            return redirect(url_for("detalhes", id=id))
        conn.execute("DELETE FROM demandas WHERE id = ?", (id,))
        conn.commit()
    finally:
        conn.close()
    flash("Demanda deletada.")
    return redirect(url_for(destino))


@app.route("/buscar")
@login_required
def buscar():
    termo_busca = request.args.get("q", "").strip()
    filtro_prioridade = request.args.get("prioridade", "").strip()
    filtro_usuario_id = request.args.get("usuario_id", "").strip() or None
    ordenacao = request.args.get("ordenacao", "prioridade")

    if not termo_busca:
        return redirect(url_for(
            "index",
            prioridade=normalize_priority(filtro_prioridade) or None,
            usuario_id=filtro_usuario_id or None,
            ordenacao=normalize_order(ordenacao),
        ))

    contexto = build_demand_context(
        STATUS_ABERTA,
        filtro_prioridade=filtro_prioridade,
        filtro_usuario_id=filtro_usuario_id,
        termo_busca=termo_busca,
        ordenacao=ordenacao,
    )
    contexto["rota_listagem"] = "buscar"
    return render_template("index.html", **contexto)


@app.route("/detalhes/<int:id>")
@login_required
def detalhes(id):
    conn = get_db()
    try:
        demanda = conn.execute(
            "SELECT d.*, u.username AS usuario_username, "
            "COALESCE(resp.nome, '') AS responsavel_nome "
            "FROM demandas d "
            "JOIN usuarios u ON u.id = d.usuario_id "
            "LEFT JOIN usuarios resp ON resp.id = d.responsavel_id "
            "WHERE d.id = ?",
            (id,),
        ).fetchone()
        if not demanda:
            abort(404)

        comentarios = conn.execute(
            "SELECT * FROM comentarios WHERE demanda_id = ? ORDER BY data DESC",
            (id,),
        ).fetchall()

        historico = conn.execute(
            "SELECT * FROM historico_status WHERE demanda_id = ? ORDER BY data ASC",
            (id,),
        ).fetchall()
    finally:
        conn.close()

    rota_voltar = "concluidas" if demanda["status"] in (STATUS_CONCLUIDA, STATUS_CANCELADA) else "index"
    return render_template(
        "detalhes.html",
        demanda=demanda,
        comentarios=comentarios,
        comentarios_total=len(comentarios),
        historico=historico,
        rota_voltar=rota_voltar,
        eh_solicitante=(session.get("usuario_id") == demanda["usuario_id"]),
    )


@app.route("/adicionar_comentario/<int:demanda_id>", methods=["POST"])
@login_required
def adicionar_comentario(demanda_id):
    _validate_csrf()
    conn = get_db()
    try:
        demanda = conn.execute("SELECT id FROM demandas WHERE id = ?", (demanda_id,)).fetchone()
        if not demanda:
            abort(404)

        comentario = request.form.get("comentario", "").strip()
        if not comentario:
            flash("O comentário não pode estar vazio.")
            return redirect(url_for("detalhes", id=demanda_id))

        conn.execute(
            "INSERT INTO comentarios (demanda_id, comentario, autor, data) VALUES (?, ?, ?, ?)",
            (
                demanda_id,
                comentario,
                session["usuario_nome"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
    finally:
        conn.close()

    flash("Comentário adicionado.")
    return redirect(url_for("detalhes", id=demanda_id))


@app.route("/usuarios")
@login_required
def usuarios():
    stats = get_all_users_with_stats()
    return render_template("usuarios.html", usuarios_stats=stats)


@app.route("/api/alerts/count")
@login_required
def api_alerts_count():
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT COUNT(*) as count FROM demandas
            WHERE prioridade = 'Crítica'
              AND status NOT IN ('Concluída', 'Cancelada')
              AND data_prevista IS NOT NULL
              AND data_prevista < datetime('now', 'localtime')
            """
        ).fetchone()
        return jsonify({"count": row["count"]})
    finally:
        conn.close()


# ── Dashboard ─────────────────────────────────────────────────────────────────

def _build_dashboard_filters():
    """Build WHERE clause and params list from dashboard filter request args.

    Returns (where_str, params_list, filter_labels_list).
    All user-supplied values are bound via ? placeholders — never interpolated.
    """
    clauses, params, labels = [], [], []

    periodo = request.args.get("periodo", "all")
    if periodo == "hoje":
        clauses.append("date(d.data_criacao) = date('now', 'localtime')")
        labels.append("Hoje")
    elif periodo == "7d":
        clauses.append("d.data_criacao >= datetime('now', '-7 days', 'localtime')")
        labels.append("Últimos 7 dias")
    elif periodo == "30d":
        clauses.append("d.data_criacao >= datetime('now', '-30 days', 'localtime')")
        labels.append("Último mês")
    elif periodo == "custom":
        d_ini = request.args.get("data_inicio", "").strip()
        d_fim = request.args.get("data_fim", "").strip()
        if d_ini:
            clauses.append("d.data_criacao >= ?")
            params.append(d_ini + " 00:00:00")
            labels.append(f"De {d_ini}")
        if d_fim:
            clauses.append("d.data_criacao <= ?")
            params.append(d_fim + " 23:59:59")
            labels.append(f"Até {d_fim}")

    uids = [v for v in request.args.getlist("usuario_id") if v.strip()]
    valid_uids = []
    for v in uids:
        try:
            valid_uids.append(int(v))
        except (ValueError, TypeError):
            pass
    if valid_uids:
        placeholders = ",".join("?" * len(valid_uids))
        clauses.append(f"d.responsavel_id IN ({placeholders})")
        params.extend(valid_uids)
        labels.append(f"Responsável ({len(valid_uids)} selecionado{'s' if len(valid_uids) > 1 else ''})")

    prioridade = request.args.get("prioridade", "").strip()
    if prioridade and prioridade in PRIORIDADES:
        clauses.append("d.prioridade = ?")
        params.append(prioridade)
        labels.append(f"Prioridade: {prioridade}")

    status = request.args.get("status", "").strip()
    if status == "Indefinido":
        clauses.append("(d.status IS NULL OR TRIM(d.status) = '')")
        labels.append("Status: Indefinido")
    elif status and status in TODOS_STATUS:
        clauses.append("d.status = ?")
        params.append(status)
        labels.append(f"Status: {status}")

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params, labels


@app.route("/")
@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db()
    try:
        usuarios = conn.execute(
            "SELECT id, nome FROM usuarios ORDER BY nome"
        ).fetchall()
    finally:
        conn.close()
    return render_template(
        "dashboard.html",
        usuarios=usuarios,
        prioridades=PRIORIDADES,
        todos_status=TODOS_STATUS,
    )


@app.route("/api/dashboard/kpis")
@login_required
def api_dashboard_kpis():
    where, params, _ = _build_dashboard_filters()
    conn = get_db()
    try:
        row = conn.execute(
            f"""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN d.status = 'Aberta' THEN 1 ELSE 0 END) as abertas,
                SUM(CASE WHEN d.status = 'Em andamento' THEN 1 ELSE 0 END) as em_andamento,
                SUM(CASE WHEN d.status = 'Concluída' THEN 1 ELSE 0 END) as concluidas,
                SUM(CASE WHEN d.status = 'Cancelada' THEN 1 ELSE 0 END) as canceladas,
                SUM(CASE WHEN d.status NOT IN ('Concluída','Cancelada')
                              AND d.data_prevista IS NOT NULL
                              AND d.data_prevista < datetime('now','localtime')
                         THEN 1 ELSE 0 END) as atrasadas,
                SUM(CASE WHEN d.prioridade = 'Crítica' THEN 1 ELSE 0 END) as criticas,
                SUM(CASE WHEN d.prioridade = 'Crítica'
                              AND d.status NOT IN ('Concluída','Cancelada')
                              AND d.data_prevista IS NOT NULL
                              AND d.data_prevista < datetime('now','localtime')
                         THEN 1 ELSE 0 END) as criticas_atrasadas,
                CASE
                    WHEN SUM(CASE WHEN d.status='Concluída' AND d.data_conclusao IS NOT NULL
                                  THEN CASE WHEN d.prioridade='Crítica' THEN 2 ELSE 1 END END) > 0
                    THEN ROUND(
                        SUM(CASE WHEN d.status='Concluída' AND d.data_conclusao IS NOT NULL
                                 THEN (julianday(d.data_conclusao) - julianday(d.data_criacao)) *
                                      CASE WHEN d.prioridade='Crítica' THEN 2 ELSE 1 END END) /
                        SUM(CASE WHEN d.status='Concluída' AND d.data_conclusao IS NOT NULL
                                 THEN CASE WHEN d.prioridade='Crítica' THEN 2 ELSE 1 END END)
                    , 1)
                    ELSE NULL
                END as tempo_medio_dias
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            """,
            params,
        ).fetchone()

        por_resp = conn.execute(
            f"""
            SELECT
                COALESCE(u.nome, 'Sem responsável') as nome,
                COUNT(d.id) as total,
                SUM(CASE WHEN d.status IN ('Aberta','Em andamento') THEN 1 ELSE 0 END) as abertas,
                SUM(CASE WHEN d.status NOT IN ('Concluída','Cancelada')
                              AND d.data_prevista IS NOT NULL
                              AND d.data_prevista < datetime('now','localtime')
                         THEN 1 ELSE 0 END) as atrasadas,
                SUM(CASE WHEN d.prioridade = 'Crítica' THEN 1 ELSE 0 END) as criticas
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            GROUP BY d.usuario_id
            ORDER BY abertas DESC, atrasadas DESC
            """,
            params,
        ).fetchall()

        total = row["total"] or 0
        return jsonify({
            "total": total,
            "abertas": row["abertas"] or 0,
            "em_andamento": row["em_andamento"] or 0,
            "concluidas": row["concluidas"] or 0,
            "canceladas": row["canceladas"] or 0,
            "atrasadas": row["atrasadas"] or 0,
            "criticas": row["criticas"] or 0,
            "criticas_atrasadas": row["criticas_atrasadas"] or 0,
            "tempo_medio_dias": row["tempo_medio_dias"],
            "pct_abertas": round((row["abertas"] or 0) / total * 100) if total else 0,
            "pct_concluidas": round((row["concluidas"] or 0) / total * 100) if total else 0,
            "pct_atrasadas": round((row["atrasadas"] or 0) / total * 100) if total else 0,
            "pct_criticas": round((row["criticas"] or 0) / total * 100) if total else 0,
            "por_responsavel": [dict(r) for r in por_resp],
        })
    finally:
        conn.close()


@app.route("/api/dashboard/charts")
@login_required
def api_dashboard_charts():
    where, params, _ = _build_dashboard_filters()
    conn = get_db()
    try:
        por_status = conn.execute(
            f"""
            SELECT COALESCE(d.status, 'Indefinido') as label, COUNT(*) as total
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            GROUP BY d.status ORDER BY total DESC
            """,
            params,
        ).fetchall()

        por_prioridade = conn.execute(
            f"""
            SELECT COALESCE(d.prioridade, 'Indefinido') as label, COUNT(*) as total
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            GROUP BY d.prioridade
            ORDER BY CASE d.prioridade
                WHEN 'Crítica' THEN 1 WHEN 'Alta' THEN 2
                WHEN 'Média' THEN 3 WHEN 'Baixa' THEN 4 ELSE 5 END
            """,
            params,
        ).fetchall()

        gran = request.args.get("granularity", "mensal")
        if gran == "diario":
            fmt, date_filter, limit = "%Y-%m-%d", "AND data_criacao >= datetime('now','-30 days','localtime')", 30
        elif gran == "semanal":
            fmt, date_filter, limit = "%Y-%W", "AND data_criacao >= datetime('now','-84 days','localtime')", 12
        else:
            fmt, date_filter, limit = "%Y-%m", "", 18

        evolucao = conn.execute(
            f"""
            SELECT
                strftime('{fmt}', data_criacao) as periodo,
                COUNT(*) as criadas,
                SUM(CASE WHEN status = 'Concluída' THEN 1 ELSE 0 END) as concluidas
            FROM demandas
            WHERE data_criacao IS NOT NULL {date_filter}
            GROUP BY periodo
            ORDER BY periodo
            LIMIT {limit}
            """,
        ).fetchall()

        return jsonify({
            "por_status": [{"label": r["label"], "total": r["total"]} for r in por_status],
            "por_prioridade": [{"label": r["label"], "total": r["total"]} for r in por_prioridade],
            "evolucao": [
                {"periodo": r["periodo"], "criadas": r["criadas"], "concluidas": r["concluidas"]}
                for r in evolucao
            ],
        })
    finally:
        conn.close()


@app.route("/api/dashboard/critical-overdue")
@login_required
def api_dashboard_critical_overdue():
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT
                d.id,
                d.titulo,
                COALESCE(u.nome, 'Não atribuído') as responsavel,
                COALESCE(sol.nome, 'Desconhecido') as solicitante,
                d.status,
                d.data_prevista,
                CAST(julianday('now','localtime') - julianday(d.data_prevista) AS INTEGER)
                    as dias_atrasados
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            LEFT JOIN usuarios sol ON sol.id = d.usuario_id
            WHERE d.prioridade = 'Crítica'
              AND d.status NOT IN ('Concluída', 'Cancelada')
              AND d.data_prevista IS NOT NULL
              AND d.data_prevista < datetime('now', 'localtime')
            ORDER BY dias_atrasados DESC, u.nome IS NULL DESC
            """,
        ).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


@app.route("/api/dashboard/data")
@login_required
def api_dashboard_data():
    where, params, _ = _build_dashboard_filters()
    gran = request.args.get("granularity", "mensal")
    conn = get_db()
    try:
        row = conn.execute(
            f"""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN d.status = 'Aberta' THEN 1 ELSE 0 END) as abertas,
                SUM(CASE WHEN d.status = 'Em andamento' THEN 1 ELSE 0 END) as em_andamento,
                SUM(CASE WHEN d.status = 'Concluída' THEN 1 ELSE 0 END) as concluidas,
                SUM(CASE WHEN d.status = 'Cancelada' THEN 1 ELSE 0 END) as canceladas,
                SUM(CASE WHEN d.status NOT IN ('Concluída','Cancelada')
                              AND d.data_prevista IS NOT NULL
                              AND d.data_prevista < datetime('now','localtime')
                         THEN 1 ELSE 0 END) as atrasadas,
                SUM(CASE WHEN d.prioridade = 'Crítica' THEN 1 ELSE 0 END) as criticas,
                SUM(CASE WHEN d.prioridade = 'Crítica'
                              AND d.status NOT IN ('Concluída','Cancelada')
                              AND d.data_prevista IS NOT NULL
                              AND d.data_prevista < datetime('now','localtime')
                         THEN 1 ELSE 0 END) as criticas_atrasadas,
                CASE
                    WHEN SUM(CASE WHEN d.status='Concluída' AND d.data_conclusao IS NOT NULL
                                  THEN CASE WHEN d.prioridade='Crítica' THEN 2 ELSE 1 END END) > 0
                    THEN ROUND(
                        SUM(CASE WHEN d.status='Concluída' AND d.data_conclusao IS NOT NULL
                                 THEN (julianday(d.data_conclusao) - julianday(d.data_criacao)) *
                                      CASE WHEN d.prioridade='Crítica' THEN 2 ELSE 1 END END) /
                        SUM(CASE WHEN d.status='Concluída' AND d.data_conclusao IS NOT NULL
                                 THEN CASE WHEN d.prioridade='Crítica' THEN 2 ELSE 1 END END)
                    , 1)
                    ELSE NULL
                END as tempo_medio_dias
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            """,
            params,
        ).fetchone()

        por_resp = conn.execute(
            f"""
            SELECT
                COALESCE(u.nome, 'Não atribuído') as nome,
                COUNT(d.id) as total,
                SUM(CASE WHEN d.status IN ('Aberta','Em andamento') THEN 1 ELSE 0 END) as abertas,
                SUM(CASE WHEN d.status NOT IN ('Concluída','Cancelada')
                              AND d.data_prevista IS NOT NULL
                              AND d.data_prevista < datetime('now','localtime')
                         THEN 1 ELSE 0 END) as atrasadas,
                SUM(CASE WHEN d.prioridade = 'Crítica' THEN 1 ELSE 0 END) as criticas
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            GROUP BY d.responsavel_id
            ORDER BY abertas DESC, atrasadas DESC
            """,
            params,
        ).fetchall()

        por_status = conn.execute(
            f"""
            SELECT COALESCE(d.status, 'Indefinido') as label, COUNT(*) as total
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            GROUP BY d.status ORDER BY total DESC
            """,
            params,
        ).fetchall()

        por_prioridade = conn.execute(
            f"""
            SELECT COALESCE(d.prioridade, 'Indefinido') as label, COUNT(*) as total
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            GROUP BY d.prioridade
            ORDER BY CASE d.prioridade
                WHEN 'Crítica' THEN 1 WHEN 'Alta' THEN 2
                WHEN 'Média' THEN 3 WHEN 'Baixa' THEN 4 ELSE 5 END
            """,
            params,
        ).fetchall()

        if gran == "diario":
            fmt, date_filter, limit = "%Y-%m-%d", "AND data_criacao >= datetime('now','-30 days','localtime')", 30
        elif gran == "semanal":
            fmt, date_filter, limit = "%Y-%W", "AND data_criacao >= datetime('now','-84 days','localtime')", 12
        else:
            fmt, date_filter, limit = "%Y-%m", "", 18

        evolucao = conn.execute(
            f"""
            SELECT
                strftime('{fmt}', data_criacao) as periodo,
                COUNT(*) as criadas,
                SUM(CASE WHEN status = 'Concluída' THEN 1 ELSE 0 END) as concluidas
            FROM demandas
            WHERE data_criacao IS NOT NULL {date_filter}
            GROUP BY periodo
            ORDER BY periodo
            LIMIT {limit}
            """,
        ).fetchall()

        critical = conn.execute(
            """
            SELECT
                d.id,
                d.titulo,
                COALESCE(u.nome, 'Sem responsável') as responsavel,
                COALESCE(sol.nome, 'Desconhecido') as solicitante,
                d.status,
                d.data_prevista,
                CAST(julianday('now','localtime') - julianday(d.data_prevista) AS INTEGER)
                    as dias_atrasados
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            LEFT JOIN usuarios sol ON sol.id = d.usuario_id
            WHERE d.prioridade = 'Crítica'
              AND d.status NOT IN ('Concluída', 'Cancelada')
              AND d.data_prevista IS NOT NULL
              AND d.data_prevista < datetime('now', 'localtime')
            ORDER BY dias_atrasados DESC, u.nome IS NULL DESC
            """,
        ).fetchall()

        total = row["total"] or 0
        return jsonify({
            "kpis": {
                "total": total,
                "abertas": row["abertas"] or 0,
                "em_andamento": row["em_andamento"] or 0,
                "concluidas": row["concluidas"] or 0,
                "canceladas": row["canceladas"] or 0,
                "atrasadas": row["atrasadas"] or 0,
                "criticas": row["criticas"] or 0,
                "criticas_atrasadas": row["criticas_atrasadas"] or 0,
                "tempo_medio_dias": row["tempo_medio_dias"],
                "pct_abertas": round((row["abertas"] or 0) / total * 100) if total else 0,
                "pct_concluidas": round((row["concluidas"] or 0) / total * 100) if total else 0,
                "pct_atrasadas": round((row["atrasadas"] or 0) / total * 100) if total else 0,
                "pct_criticas": round((row["criticas"] or 0) / total * 100) if total else 0,
                "por_responsavel": [dict(r) for r in por_resp],
            },
            "charts": {
                "por_status": [{"label": r["label"], "total": r["total"]} for r in por_status],
                "por_prioridade": [{"label": r["label"], "total": r["total"]} for r in por_prioridade],
                "evolucao": [
                    {"periodo": r["periodo"], "criadas": r["criadas"], "concluidas": r["concluidas"]}
                    for r in evolucao
                ],
            },
            "critical": [dict(r) for r in critical],
        })
    finally:
        conn.close()


@app.route("/api/dashboard/export")
@login_required
def api_dashboard_export():
    tipo = request.args.get("type", "xlsx")
    where, params, filter_labels = _build_dashboard_filters()

    conn = get_db()
    try:
        rows = conn.execute(
            f"""
            SELECT
                d.id,
                d.titulo,
                COALESCE(u.nome, 'Não atribuído') as responsavel,
                d.prioridade,
                d.status,
                d.data_criacao,
                COALESCE(d.data_prevista, '') as data_prevista,
                COALESCE(d.data_conclusao, '') as data_conclusao,
                CASE WHEN d.status NOT IN ('Concluída','Cancelada')
                          AND d.data_prevista IS NOT NULL
                          AND d.data_prevista < datetime('now','localtime')
                     THEN 'Sim' ELSE 'Não' END as atrasada
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            {where}
            ORDER BY
                CASE d.prioridade WHEN 'Crítica' THEN 1 WHEN 'Alta' THEN 2
                                  WHEN 'Média' THEN 3 ELSE 4 END,
                d.data_criacao DESC
            """,
            params,
        ).fetchall()
    finally:
        conn.close()

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    filter_desc = " · ".join(filter_labels) if filter_labels else "Sem filtros ativos"
    headers_row = [
        "ID", "Título", "Responsável", "Prioridade", "Status",
        "Criado em", "Previsto para", "Concluído em", "Atrasada",
    ]

    if tipo == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["SGDI — Sistema de Gestão de Demandas"])
        writer.writerow(["Dashboard Gerencial de Demandas"])
        writer.writerow([f"Gerado em: {now_str}"])
        writer.writerow([])
        writer.writerow(headers_row)
        for r in rows:
            writer.writerow([
                r["id"], r["titulo"], r["responsavel"], r["prioridade"],
                r["status"], r["data_criacao"], r["data_prevista"],
                r["data_conclusao"], r["atrasada"],
            ])
        writer.writerow([])
        writer.writerow([f"Filtros aplicados: {filter_desc}"])
        buf.seek(0)
        return Response(
            buf.getvalue().encode("utf-8-sig"),
            mimetype="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment; filename=demandas_dashboard.csv"},
        )

    if tipo == "pdf":
        return _export_pdf(rows, now_str, filter_desc, filter_labels)

    # Excel export (default)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl não instalado. Execute: pip install openpyxl"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Demandas"

    ws.merge_cells("A1:I1")
    ws["A1"] = "SGDI — Sistema de Gestão de Demandas"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = "Dashboard Gerencial de Demandas"
    ws["A2"].font = Font(size=11, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:I3")
    ws["A3"] = f"Gerado em: {now_str}"
    ws["A3"].font = Font(size=10, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.append([])

    ws.append(headers_row)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col_idx in range(1, len(headers_row) + 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")

    priority_fills = {
        "Crítica": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Alta":    PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Média":   PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "Baixa":   PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    }
    status_fills = {
        "Aberta":        PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "Em andamento":  PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Concluída":     PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Cancelada":     PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
    }

    for r in rows:
        ws.append([
            r["id"], r["titulo"], r["responsavel"], r["prioridade"],
            r["status"], r["data_criacao"], r["data_prevista"],
            r["data_conclusao"], r["atrasada"],
        ])
        rn = ws.max_row
        if r["prioridade"] in priority_fills:
            ws.cell(row=rn, column=4).fill = priority_fills[r["prioridade"]]
        if r["status"] in status_fills:
            ws.cell(row=rn, column=5).fill = status_fills[r["status"]]
        if r["atrasada"] == "Sim":
            ws.cell(row=rn, column=9).font = Font(bold=True, color="DC2626")

    ws.append([])
    ws.merge_cells(f"A{ws.max_row + 1}:I{ws.max_row + 1}")
    ws.append([f"Filtros aplicados: {filter_desc}"])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells(f"A{ws.max_row}:I{ws.max_row}")

    for idx, width in enumerate([6, 42, 22, 12, 15, 20, 20, 20, 10], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="demandas_dashboard.xlsx",
    )


@app.route("/api/dashboard/critical-overdue/export")
@login_required
def api_dashboard_critical_overdue_export():
    tipo = request.args.get("type", "xlsx")
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT
                d.id,
                d.titulo,
                COALESCE(u.nome, 'Não atribuído') as responsavel,
                COALESCE(sol.nome, 'Desconhecido') as solicitante,
                d.status,
                d.data_criacao,
                d.data_prevista,
                CAST(julianday('now','localtime') - julianday(d.data_prevista) AS INTEGER)
                    as dias_atrasados
            FROM demandas d
            LEFT JOIN usuarios u ON u.id = d.responsavel_id
            LEFT JOIN usuarios sol ON sol.id = d.usuario_id
            WHERE d.prioridade = 'Crítica'
              AND d.status NOT IN ('Concluída', 'Cancelada')
              AND d.data_prevista IS NOT NULL
              AND d.data_prevista < datetime('now', 'localtime')
            ORDER BY dias_atrasados DESC, u.nome IS NULL DESC
            """,
        ).fetchall()
    finally:
        conn.close()

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    headers_row = ["ID", "Título", "Responsável", "Solicitante", "Status", "Criado em", "SLA Previsto", "Dias Atrasados"]

    if tipo == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["SGDI — Relatório: Críticas e Atrasadas"])
        writer.writerow([f"Gerado em: {now_str} · Prioridade CRÍTICA com SLA vencido"])
        writer.writerow([])
        writer.writerow(headers_row)
        for r in rows:
            writer.writerow([
                r["id"], r["titulo"], r["responsavel"], r["solicitante"], r["status"],
                r["data_criacao"], r["data_prevista"], str(r["dias_atrasados"]) + "d",
            ])
        buf.seek(0)
        return Response(
            buf.getvalue().encode("utf-8-sig"),
            mimetype="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment; filename=criticas_atrasadas.csv"},
        )

    if tipo == "pdf":
        return _export_critical_pdf(rows, now_str)

    # Excel (default)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl não instalado. Execute: pip install openpyxl"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Críticas e Atrasadas"

    ws.merge_cells("A1:H1")
    ws["A1"] = "SGDI — Relatório: Críticas e Atrasadas"
    ws["A1"].font = Font(bold=True, size=14, color="DC2626")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Gerado em: {now_str} · Prioridade CRÍTICA · SLA vencido"
    ws["A2"].font = Font(size=10, color="B91C1C")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])

    ws.append(headers_row)
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    for col_idx in range(1, len(headers_row) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    for i, r in enumerate(rows):
        ws.append([
            r["id"], r["titulo"], r["responsavel"], r["solicitante"], r["status"],
            r["data_criacao"], r["data_prevista"], r["dias_atrasados"],
        ])
        rn = ws.max_row
        ws.cell(row=rn, column=8).font = Font(bold=True, color="DC2626")
        if i % 2 == 1:
            for col in range(1, 9):
                ws.cell(row=rn, column=col).fill = alt_fill

    for idx, width in enumerate([6, 42, 22, 22, 15, 20, 20, 16], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="criticas_atrasadas.xlsx",
    )


def _export_critical_pdf(rows, now_str):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        return jsonify({"error": "reportlab não instalado"}), 500

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    base = getSampleStyleSheet()

    def _ps(name, **kw):
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    s_title  = _ps("CrTitle",  fontSize=16, fontName="Helvetica-Bold", spaceAfter=4,
                   textColor=colors.HexColor("#DC2626"), alignment=TA_CENTER)
    s_sub    = _ps("CrSub",    fontSize=10, spaceAfter=14,
                   textColor=colors.HexColor("#B91C1C"), alignment=TA_CENTER)
    s_cell   = _ps("CrCell",   fontSize=7.5, leading=10, textColor=colors.HexColor("#0f172a"))
    s_cell_c = _ps("CrCellC",  fontSize=7.5, leading=10, textColor=colors.HexColor("#0f172a"),
                   alignment=TA_CENTER)
    s_hdr    = _ps("CrHdr",    fontSize=7.5, fontName="Helvetica-Bold",
                   textColor=colors.white, alignment=TA_CENTER)
    s_days   = _ps("CrDays",   fontSize=8, fontName="Helvetica-Bold",
                   textColor=colors.HexColor("#DC2626"), alignment=TA_CENTER)

    header_labels = ["ID", "Título", "Responsável", "Status", "Criado em", "SLA Previsto", "Dias Atrasados"]
    tbl_data = [[Paragraph(h, s_hdr) for h in header_labels]]
    for r in rows:
        dc = (r["data_criacao"] or "")[:10]
        dp = (r["data_prevista"] or "—")[:10]
        tbl_data.append([
            Paragraph(str(r["id"]),          s_cell_c),
            Paragraph(r["titulo"],           s_cell),
            Paragraph(r["responsavel"],      s_cell),
            Paragraph(r["status"],           s_cell_c),
            Paragraph(dc,                    s_cell_c),
            Paragraph(dp,                    s_cell_c),
            Paragraph(str(r["dias_atrasados"]) + "d", s_days),
        ])

    col_w = [0.9*cm, 8*cm, 4*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm]
    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)

    ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#DC2626")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#FECACA")),
        ("LINEBELOW",     (0, 0), (-1, 0), 1, colors.HexColor("#B91C1C")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
    ])
    for i in range(1, len(tbl_data)):
        if i % 2 == 0:
            ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFF5F5"))
    tbl.setStyle(ts)

    doc.build([
        Paragraph("SGDI — Relatório: Críticas e Atrasadas", s_title),
        Paragraph(f"Gerado em: {now_str} · Prioridade CRÍTICA com SLA vencido", s_sub),
        Spacer(1, 0.2 * cm),
        tbl,
    ])
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="criticas_atrasadas.pdf",
    )


def _export_pdf(rows, now_str, filter_desc, filter_labels):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        return jsonify({"error": "reportlab não instalado"}), 500

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    base = getSampleStyleSheet()

    def _ps(name, **kw):
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    s_title    = _ps("PdfTitle",    fontSize=16, fontName="Helvetica-Bold", spaceAfter=4,
                     textColor=colors.HexColor("#0f172a"), alignment=TA_CENTER)
    s_subtitle = _ps("PdfSub",     fontSize=11, spaceAfter=4,
                     textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)
    s_meta     = _ps("PdfMeta",    fontSize=9, spaceAfter=14,
                     textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    s_footer   = _ps("PdfFooter",  fontSize=8, textColor=colors.HexColor("#94a3b8"))
    s_cell     = _ps("PdfCell",    fontSize=7.5, leading=10, textColor=colors.HexColor("#0f172a"))
    s_cell_c   = _ps("PdfCellC",   fontSize=7.5, leading=10, textColor=colors.HexColor("#0f172a"),
                     alignment=TA_CENTER)
    s_hdr      = _ps("PdfHdr",     fontSize=7.5, fontName="Helvetica-Bold",
                     textColor=colors.white, alignment=TA_CENTER)
    s_late     = _ps("PdfLate",    fontSize=7.5, leading=10, fontName="Helvetica-Bold",
                     textColor=colors.HexColor("#DC2626"), alignment=TA_CENTER)

    header_labels = [
        "ID", "Título", "Responsável", "Prioridade", "Status",
        "Criado em", "Previsto para", "Concluído em", "Atrasada",
    ]

    tbl_data = [[Paragraph(h, s_hdr) for h in header_labels]]
    for r in rows:
        dc = (r["data_criacao"] or "")[:10]
        dp = (r["data_prevista"] or "—")[:10]
        dk = (r["data_conclusao"] or "—")[:10]
        tbl_data.append([
            Paragraph(str(r["id"]),       s_cell_c),
            Paragraph(r["titulo"],        s_cell),
            Paragraph(r["responsavel"],   s_cell),
            Paragraph(r["prioridade"],    s_cell_c),
            Paragraph(r["status"],        s_cell_c),
            Paragraph(dc,                 s_cell_c),
            Paragraph(dp,                 s_cell_c),
            Paragraph(dk,                 s_cell_c),
            Paragraph(r["atrasada"],      s_late if r["atrasada"] == "Sim" else s_cell_c),
        ])

    # Landscape A4 usable width ≈ 26.7 cm
    col_w = [0.9*cm, 7.8*cm, 3.5*cm, 2*cm, 2.8*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.2*cm]

    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)

    prio_bg = {
        "Crítica": colors.HexColor("#FEE2E2"),
        "Alta":    colors.HexColor("#FEF3C7"),
        "Média":   colors.HexColor("#DBEAFE"),
        "Baixa":   colors.HexColor("#D1FAE5"),
    }
    stat_bg = {
        "Aberta":       colors.HexColor("#DBEAFE"),
        "Em andamento": colors.HexColor("#FEF3C7"),
        "Concluída":    colors.HexColor("#D1FAE5"),
        "Cancelada":    colors.HexColor("#F1F5F9"),
    }

    ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("LINEBELOW",     (0, 0), (-1, 0), 1, colors.HexColor("#1d4ed8")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
    ])

    for i in range(1, len(tbl_data)):
        if i % 2 == 0:
            ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAFC"))
        r = rows[i - 1]
        if r["prioridade"] in prio_bg:
            ts.add("BACKGROUND", (3, i), (3, i), prio_bg[r["prioridade"]])
        if r["status"] in stat_bg:
            ts.add("BACKGROUND", (4, i), (4, i), stat_bg[r["status"]])

    tbl.setStyle(ts)

    story = [
        Paragraph("SGDI — Sistema de Gestão de Demandas", s_title),
        Paragraph("Dashboard Gerencial de Demandas", s_subtitle),
        Paragraph(f"Gerado em: {now_str}", s_meta),
        Spacer(1, 0.2 * cm),
        tbl,
    ]
    if filter_labels:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph(f"Filtros aplicados: {filter_desc}", s_footer))

    doc.build(story)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="demandas_dashboard.pdf",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Gestão de API Keys (interface web — login obrigatório)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/keys", methods=["GET", "POST"])
@login_required
def api_keys():
    conn = get_db()
    try:
        if request.method == "POST":
            _validate_csrf()
            acao = request.form.get("acao")

            if acao == "criar":
                descricao = request.form.get("descricao", "").strip()
                if not descricao:
                    flash("Informe uma descrição para a chave.")
                    return redirect(url_for("api_keys"))
                nova_chave = secrets.token_urlsafe(32)
                conn.execute(
                    "INSERT INTO api_keys (chave, descricao, criado_por) VALUES (?, ?, ?)",
                    (nova_chave, descricao, session["usuario_id"]),
                )
                conn.commit()
                flash(f"Chave gerada: {nova_chave} — copie agora, não será exibida novamente.")

            elif acao == "revogar":
                key_id = request.form.get("key_id")
                conn.execute("UPDATE api_keys SET ativo = 0 WHERE id = ?", (key_id,))
                conn.commit()
                flash("Chave revogada com sucesso.")

            return redirect(url_for("api_keys"))

        keys = conn.execute(
            """
            SELECT k.id, k.descricao, k.ativo, k.criado_em,
                   SUBSTR(k.chave, 1, 8) || '••••••••••••••••••••••••••••••' as chave_mascarada,
                   u.nome as criado_por_nome
            FROM api_keys k
            LEFT JOIN usuarios u ON u.id = k.criado_por
            ORDER BY k.criado_em DESC
            """
        ).fetchall()
        return render_template("api_keys.html", keys=keys)
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# API REST Externa v1 — autenticação por X-API-Key
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/v1/demandas", methods=["GET"])
@api_key_required
def api_v1_demandas_list():
    """
    Lista todas as demandas com filtros opcionais.
    ---
    tags: [Demandas]
    parameters:
      - name: status
        in: query
        type: string
        enum: [Aberta, Em andamento, Concluída, Cancelada]
        description: Filtrar por status
      - name: prioridade
        in: query
        type: string
        enum: [Crítica, Alta, Média, Baixa]
        description: Filtrar por prioridade
      - name: responsavel_id
        in: query
        type: integer
        description: Filtrar por ID do responsável
      - name: limit
        in: query
        type: integer
        default: 50
        description: Máximo de registros retornados (máx. 200)
      - name: offset
        in: query
        type: integer
        default: 0
        description: Paginação — número de registros a pular
    responses:
      200:
        description: Lista de demandas
        schema:
          properties:
            success: {type: boolean}
            data:
              type: array
              items:
                properties:
                  id: {type: integer}
                  titulo: {type: string}
                  status: {type: string}
                  prioridade: {type: string}
                  solicitante: {type: string}
                  responsavel: {type: string}
                  data_criacao: {type: string}
                  data_prevista: {type: string}
            meta:
              properties:
                total: {type: integer}
      401:
        description: Header X-API-Key ausente
      403:
        description: Chave inválida ou desativada
    """
    status     = request.args.get("status", "").strip()
    prioridade = request.args.get("prioridade", "").strip()
    resp_id    = request.args.get("responsavel_id", "").strip()
    try:
        limit  = min(int(request.args.get("limit",  50)), 200)
        offset = max(int(request.args.get("offset",  0)),   0)
    except ValueError:
        return _api_err("Parâmetros limit/offset devem ser inteiros", 400)

    clauses, params = [], []
    if status:
        clauses.append("d.status = ?"); params.append(status)
    if prioridade:
        clauses.append("d.prioridade = ?"); params.append(prioridade)
    if resp_id:
        clauses.append("d.responsavel_id = ?"); params.append(resp_id)

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    conn = get_db()
    try:
        total = conn.execute(
            f"SELECT COUNT(*) FROM demandas d {where}", params
        ).fetchone()[0]
        rows = conn.execute(
            f"""
            SELECT d.id, d.titulo, d.descricao, d.status, d.prioridade,
                   d.solicitante, d.data_criacao, d.data_prevista, d.data_conclusao,
                   COALESCE(resp.nome, 'Não atribuído') as responsavel
            FROM demandas d
            LEFT JOIN usuarios resp ON resp.id = d.responsavel_id
            {where}
            ORDER BY {PRIORIDADE_ORDEM_SQL}, d.data_criacao DESC
            LIMIT ? OFFSET ?
            """,
            params + [limit, offset],
        ).fetchall()
        return _api_ok([dict(r) for r in rows], total=total)
    finally:
        conn.close()


@app.route("/api/v1/demandas", methods=["POST"])
@api_key_required
def api_v1_demandas_create():
    """
    Cria uma nova demanda.
    ---
    tags: [Demandas]
    consumes: [application/json]
    parameters:
      - in: body
        name: body
        required: true
        schema:
          required: [titulo, descricao, solicitante, prioridade]
          properties:
            titulo:
              type: string
              example: Falha no módulo de relatórios
            descricao:
              type: string
              example: O módulo trava ao gerar relatórios com mais de 1000 linhas.
            solicitante:
              type: string
              example: Sistema ERP
            prioridade:
              type: string
              enum: [Crítica, Alta, Média, Baixa]
              example: Alta
            responsavel_id:
              type: integer
              example: 2
            data_prevista:
              type: string
              example: "2026-06-30"
    responses:
      201:
        description: Demanda criada com sucesso
      400:
        description: Dados inválidos ou campos obrigatórios ausentes
    """
    body = request.get_json(silent=True) or {}
    titulo      = (body.get("titulo")     or "").strip()
    descricao   = (body.get("descricao")  or "").strip()
    solicitante = (body.get("solicitante") or "").strip()
    prioridade  = (body.get("prioridade") or "").strip()
    responsavel_id  = body.get("responsavel_id")
    data_prevista   = (body.get("data_prevista") or "").strip() or None

    if not all([titulo, descricao, solicitante, prioridade]):
        return _api_err("Campos obrigatórios: titulo, descricao, solicitante, prioridade", 400)
    if prioridade not in PRIORIDADES:
        return _api_err(f"prioridade inválida. Use: {', '.join(PRIORIDADES)}", 400)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        cursor = conn.execute(
            """INSERT INTO demandas
               (titulo, descricao, solicitante, usuario_id, prioridade,
                status, data_criacao, data_prevista, responsavel_id)
               VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?)""",
            (titulo, descricao, solicitante, prioridade,
             STATUS_ABERTA, now, data_prevista, responsavel_id),
        )
        demanda_id = cursor.lastrowid
        conn.execute(
            "INSERT INTO historico_status (demanda_id, status_anterior, status_novo, autor, data) "
            "VALUES (?, ?, ?, ?, ?)",
            (demanda_id, None, STATUS_ABERTA, f"API/{solicitante}", now),
        )
        conn.commit()
        return _api_ok({"id": demanda_id, "status": STATUS_ABERTA}), 201
    finally:
        conn.close()


@app.route("/api/v1/demandas/<int:id>", methods=["GET"])
@api_key_required
def api_v1_demandas_get(id):
    """
    Retorna os detalhes de uma demanda específica.
    ---
    tags: [Demandas]
    parameters:
      - name: id
        in: path
        type: integer
        required: true
        description: ID da demanda
    responses:
      200:
        description: Detalhes da demanda
      404:
        description: Demanda não encontrada
    """
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT d.id, d.titulo, d.descricao, d.status, d.prioridade,
                   d.solicitante, d.data_criacao, d.data_prevista, d.data_conclusao,
                   COALESCE(resp.nome, 'Não atribuído') as responsavel,
                   u.nome as solicitante_usuario
            FROM demandas d
            LEFT JOIN usuarios resp ON resp.id = d.responsavel_id
            LEFT JOIN usuarios u   ON u.id   = d.usuario_id
            WHERE d.id = ?
            """,
            (id,),
        ).fetchone()
        if not row:
            return _api_err("Demanda não encontrada", 404)
        return _api_ok(dict(row))
    finally:
        conn.close()


@app.route("/api/v1/demandas/<int:id>/status", methods=["PATCH"])
@api_key_required
def api_v1_demandas_status(id):
    """
    Atualiza o status de uma demanda.
    ---
    tags: [Demandas]
    consumes: [application/json]
    parameters:
      - name: id
        in: path
        type: integer
        required: true
      - in: body
        name: body
        required: true
        schema:
          required: [status]
          properties:
            status:
              type: string
              enum: [Aberta, Em andamento, Concluída, Cancelada]
              example: Em andamento
            autor:
              type: string
              example: Sistema ERP
    responses:
      200:
        description: Status atualizado com sucesso
      400:
        description: Status inválido
      404:
        description: Demanda não encontrada
    """
    body   = request.get_json(silent=True) or {}
    novo   = (body.get("status") or "").strip()
    autor  = (body.get("autor")  or "API").strip()

    if novo not in TODOS_STATUS:
        return _api_err(f"Status inválido. Use: {', '.join(TODOS_STATUS)}", 400)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        row = conn.execute("SELECT status FROM demandas WHERE id = ?", (id,)).fetchone()
        if not row:
            return _api_err("Demanda não encontrada", 404)
        anterior = row["status"]
        conclusao = now if novo == STATUS_CONCLUIDA else None
        conn.execute(
            "UPDATE demandas SET status = ?, data_conclusao = COALESCE(?, data_conclusao) WHERE id = ?",
            (novo, conclusao, id),
        )
        conn.execute(
            "INSERT INTO historico_status (demanda_id, status_anterior, status_novo, autor, data) "
            "VALUES (?, ?, ?, ?, ?)",
            (id, anterior, novo, autor, now),
        )
        conn.commit()
        return _api_ok({"id": id, "status_anterior": anterior, "status_novo": novo})
    finally:
        conn.close()


@app.route("/api/v1/demandas/<int:id>/comentarios", methods=["GET"])
@api_key_required
def api_v1_comentarios_list(id):
    """
    Lista os comentários de uma demanda.
    ---
    tags: [Comentários]
    parameters:
      - name: id
        in: path
        type: integer
        required: true
    responses:
      200:
        description: Lista de comentários
      404:
        description: Demanda não encontrada
    """
    conn = get_db()
    try:
        if not conn.execute("SELECT 1 FROM demandas WHERE id = ?", (id,)).fetchone():
            return _api_err("Demanda não encontrada", 404)
        rows = conn.execute(
            "SELECT id, autor, comentario, data FROM comentarios WHERE demanda_id = ? ORDER BY data",
            (id,),
        ).fetchall()
        return _api_ok([dict(r) for r in rows], total=len(rows))
    finally:
        conn.close()


@app.route("/api/v1/demandas/<int:id>/comentarios", methods=["POST"])
@api_key_required
def api_v1_comentarios_create(id):
    """
    Adiciona um comentário a uma demanda.
    ---
    tags: [Comentários]
    consumes: [application/json]
    parameters:
      - name: id
        in: path
        type: integer
        required: true
      - in: body
        name: body
        required: true
        schema:
          required: [autor, comentario]
          properties:
            autor:
              type: string
              example: Sistema ERP
            comentario:
              type: string
              example: Demanda recebida e em análise pelo time de integração.
    responses:
      201:
        description: Comentário adicionado
      400:
        description: Campos obrigatórios ausentes
      404:
        description: Demanda não encontrada
    """
    body = request.get_json(silent=True) or {}
    autor      = (body.get("autor")      or "").strip()
    comentario = (body.get("comentario") or "").strip()

    if not autor or not comentario:
        return _api_err("Campos obrigatórios: autor, comentario", 400)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        if not conn.execute("SELECT 1 FROM demandas WHERE id = ?", (id,)).fetchone():
            return _api_err("Demanda não encontrada", 404)
        cursor = conn.execute(
            "INSERT INTO comentarios (demanda_id, comentario, autor, data) VALUES (?, ?, ?, ?)",
            (id, comentario, autor, now),
        )
        conn.commit()
        return _api_ok({"id": cursor.lastrowid}), 201
    finally:
        conn.close()


@app.route("/api/v1/usuarios", methods=["GET"])
@api_key_required
def api_v1_usuarios_list():
    """
    Lista os usuários ativos do sistema.
    ---
    tags: [Usuários]
    responses:
      200:
        description: Lista de usuários
        schema:
          properties:
            success: {type: boolean}
            data:
              type: array
              items:
                properties:
                  id: {type: integer}
                  username: {type: string}
                  nome: {type: string}
    """
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT id, username, nome FROM usuarios ORDER BY nome"
        ).fetchall()
        return _api_ok([dict(r) for r in rows], total=len(rows))
    finally:
        conn.close()


if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(debug=debug, host="0.0.0.0")
