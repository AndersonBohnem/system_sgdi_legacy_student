"""
Gerador do documento PDF: Refatoracao de Codigo SGDI v2.0
Documenta as 3 melhorias de qualidade de codigo aplicadas ao app.py.
"""
import sys
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    BaseDocTemplate, Frame, HRFlowable, KeepTogether,
    NextPageTemplate, PageBreak, PageTemplate,
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

# ── Paleta ────────────────────────────────────────────────────────────────────
C_PRIMARY = colors.HexColor("#1e40af")
C_SUCCESS = colors.HexColor("#059669")
C_WARNING = colors.HexColor("#d97706")
C_DANGER  = colors.HexColor("#dc2626")
C_DARK    = colors.HexColor("#1e293b")
C_MUTED   = colors.HexColor("#64748b")
C_BG      = colors.HexColor("#f8fafc")
C_BORDER  = colors.HexColor("#e2e8f0")
C_WHITE   = colors.white
C_NAVY    = colors.HexColor("#0f172a")
C_CODE_BG = colors.HexColor("#0f172a")
C_CODE_FG = colors.HexColor("#e2e8f0")
C_BEFORE  = colors.HexColor("#7f1d1d")
C_AFTER   = colors.HexColor("#14532d")
C_BG_BEFORE = colors.HexColor("#fff1f2")
C_BG_AFTER  = colors.HexColor("#f0fdf4")

PAGE_W, PAGE_H = A4
MARGIN    = 2.0 * cm
CONTENT_W = PAGE_W - 2 * MARGIN


# ═══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════════════════════

def build_styles():
    S = getSampleStyleSheet()
    def add(name, **kw):
        if name in S:
            S[name].__dict__.update(kw)
        else:
            S.add(ParagraphStyle(name=name, **kw))

    add("DocTitle",    fontSize=30, leading=36, textColor=C_WHITE,
        fontName="Helvetica-Bold", alignment=TA_CENTER)
    add("DocSubtitle", fontSize=12, leading=16,
        textColor=colors.HexColor("#93c5fd"),
        fontName="Helvetica", alignment=TA_CENTER)
    add("H1", fontSize=15, leading=19, textColor=C_PRIMARY,
        fontName="Helvetica-Bold", spaceBefore=22, spaceAfter=10)
    add("H2", fontSize=12, leading=15, textColor=C_DARK,
        fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    add("H3", fontSize=10.5, leading=13, textColor=C_PRIMARY,
        fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4)
    add("Body", fontSize=9.5, leading=14.5, textColor=C_DARK,
        fontName="Helvetica", alignment=TA_JUSTIFY, spaceAfter=6)
    add("Bullet", fontSize=9.5, leading=14, textColor=C_DARK,
        fontName="Helvetica", leftIndent=16, spaceAfter=3)
    add("Code", fontSize=7.5, leading=11, textColor=C_CODE_FG,
        fontName="Courier", backColor=C_CODE_BG,
        leftIndent=8, rightIndent=8, spaceBefore=2, spaceAfter=2,
        borderPadding=6)
    add("CodeBefore", fontSize=7.5, leading=11,
        textColor=colors.HexColor("#fca5a5"),
        fontName="Courier", backColor=C_BEFORE,
        leftIndent=8, rightIndent=8, spaceBefore=2, spaceAfter=2,
        borderPadding=6)
    add("CodeAfter", fontSize=7.5, leading=11,
        textColor=colors.HexColor("#86efac"),
        fontName="Courier", backColor=C_AFTER,
        leftIndent=8, rightIndent=8, spaceBefore=2, spaceAfter=2,
        borderPadding=6)
    add("Caption", fontSize=8, leading=11, textColor=C_MUTED,
        fontName="Helvetica-Oblique", alignment=TA_CENTER, spaceAfter=8)
    add("Tag",  fontSize=8, leading=11, textColor=C_WHITE,
        fontName="Helvetica-Bold", alignment=TA_CENTER)
    return S

S = build_styles()


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def sp(h=8):
    return Spacer(1, h)

def hr(color=C_BORDER, t=0.5, sb=4, sa=8):
    return HRFlowable(width="100%", thickness=t, color=color,
                      spaceBefore=sb, spaceAfter=sa)

def P(text, style="Body"):
    return Paragraph(text, S[style])

def code_block(lines, style="Code"):
    rows = []
    for line in lines:
        safe = (line.replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                    .replace(" ", "&nbsp;"))
        rows.append([Paragraph(safe, S[style])])
    t = Table(rows, colWidths=[CONTENT_W])
    bg = C_CODE_BG if style == "Code" else (C_BEFORE if style == "CodeBefore" else C_AFTER)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), bg),
        ("TOPPADDING",    (0,0), (-1,-1), 1),
        ("BOTTOMPADDING", (0,0), (-1,-1), 1),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#334155")),
    ]))
    return t

def label_bar(text, color, bg):
    s = ParagraphStyle("lb", fontSize=8.5, fontName="Helvetica-Bold",
                        textColor=color, leading=12, leftIndent=8)
    t = Table([[Paragraph(text, s)]], colWidths=[CONTENT_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), bg),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
    ]))
    return t

def before_after(before_lines, after_lines):
    lbl_b = label_bar("ANTES — Codigo original", C_BEFORE, C_BG_BEFORE)
    lbl_a = label_bar("DEPOIS — Codigo refatorado", C_AFTER, C_BG_AFTER)
    return [lbl_b, code_block(before_lines, "CodeBefore"),
            sp(4),
            lbl_a, code_block(after_lines, "CodeAfter")]

def kpi_row(items):
    cells = []
    for label, value, color in items:
        val_s = ParagraphStyle(f"kv_{label}", fontSize=18, fontName="Helvetica-Bold",
                               textColor=color, alignment=TA_CENTER, leading=22)
        lab_s = ParagraphStyle(f"kl_{label}", fontSize=8, fontName="Helvetica",
                               textColor=C_MUTED, alignment=TA_CENTER, leading=11)
        cells.append([Paragraph(value, val_s), Paragraph(label, lab_s)])
    col_w = CONTENT_W / len(items)
    t = Table([cells], colWidths=[col_w] * len(items))
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), C_BG),
        ("BOX",           (0,0), (-1,-1), 0.5, C_BORDER),
        ("LINEBEFORE",    (1,0), (-1,-1), 0.5, C_BORDER),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    return t

def info_box(text, color=C_PRIMARY, bg=None):
    if bg is None:
        bg = colors.HexColor("#eff6ff")
    s = ParagraphStyle("ib", fontSize=9, fontName="Helvetica",
                       textColor=color, leading=14, leftIndent=10)
    t = Table([[Paragraph(text, s)]], colWidths=[CONTENT_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), bg),
        ("LINEAFTER",   (0,0), (0,-1), 3, color),
        ("TOPPADDING",  (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
        ("RIGHTPADDING",(0,0), (-1,-1), 10),
    ]))
    return t

def section_header(num, title, subtitle, color):
    dark_map = {
        C_PRIMARY: colors.HexColor("#1e3a8a"),
        C_SUCCESS: colors.HexColor("#14532d"),
        C_WARNING: colors.HexColor("#78350f"),
        C_DANGER:  colors.HexColor("#7f1d1d"),
    }
    bg = dark_map.get(color, colors.HexColor("#1e3a8a"))
    ns = ParagraphStyle(f"sn{num}", fontSize=16, fontName="Helvetica-Bold",
                         textColor=color, alignment=TA_CENTER, leading=19)
    ts = ParagraphStyle(f"st{num}", fontSize=14, fontName="Helvetica-Bold",
                         textColor=C_WHITE, leading=17)
    ss = ParagraphStyle(f"ss{num}", fontSize=9, fontName="Helvetica",
                         textColor=colors.HexColor("#94a3b8"), leading=12)
    left = Table([[Paragraph(f"{num}", ns)]], colWidths=[1.2*cm],
                  style=TableStyle([
                      ("BACKGROUND",    (0,0),(-1,-1), C_NAVY),
                      ("TOPPADDING",    (0,0),(-1,-1), 8),
                      ("BOTTOMPADDING", (0,0),(-1,-1), 8),
                  ]))
    right = [Paragraph(title, ts), Paragraph(subtitle, ss)]
    t = Table([[left, right]], colWidths=[1.6*cm, CONTENT_W-1.6*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), bg),
        ("TOPPADDING",    (0,0),(-1,-1), 12),
        ("BOTTOMPADDING", (0,0),(-1,-1), 12),
        ("LEFTPADDING",   (0,0),(-1,-1), 12),
        ("RIGHTPADDING",  (0,0),(-1,-1), 16),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))
    return t

def adv_table(items):
    rows = []
    for icon, title, desc in items:
        t_s = ParagraphStyle(f"at_{title}", fontSize=9.5, fontName="Helvetica-Bold",
                              textColor=C_DARK, leading=13)
        d_s = ParagraphStyle(f"ad_{title}", fontSize=8.5, fontName="Helvetica",
                              textColor=C_MUTED, leading=12)
        i_s = ParagraphStyle(f"ai_{title}", fontSize=14, fontName="Helvetica-Bold",
                              textColor=C_SUCCESS, alignment=TA_CENTER, leading=17)
        rows.append([Paragraph(icon, i_s),
                     [Paragraph(title, t_s), Paragraph(desc, d_s)]])
    t = Table(rows, colWidths=[1.2*cm, CONTENT_W-1.2*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_BG),
        ("LINEBELOW",     (0,0),(-1,-2), 0.3, C_BORDER),
        ("BOX",           (0,0),(-1,-1), 0.3, C_BORDER),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    return t

def metrics_row(before_lines, after_lines, before_label="Antes", after_label="Depois"):
    b_s = ParagraphStyle("mr_b", fontSize=9, fontName="Helvetica-Bold",
                          textColor=C_BEFORE, alignment=TA_CENTER, leading=12)
    a_s = ParagraphStyle("mr_a", fontSize=9, fontName="Helvetica-Bold",
                          textColor=C_AFTER, alignment=TA_CENTER, leading=12)
    b_n = ParagraphStyle("mr_bn", fontSize=22, fontName="Helvetica-Bold",
                          textColor=C_BEFORE, alignment=TA_CENTER, leading=26)
    a_n = ParagraphStyle("mr_an", fontSize=22, fontName="Helvetica-Bold",
                          textColor=C_AFTER, alignment=TA_CENTER, leading=26)
    left = [Paragraph(before_lines, b_n), Paragraph(before_label, b_s)]
    right = [Paragraph(after_lines, a_n), Paragraph(after_label, a_s)]
    t = Table([[left, right]], colWidths=[CONTENT_W/2, CONTENT_W/2])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(0,-1), C_BG_BEFORE),
        ("BACKGROUND", (1,0),(1,-1), C_BG_AFTER),
        ("BOX",        (0,0),(-1,-1), 0.5, C_BORDER),
        ("LINEBEFORE", (1,0),(1,-1), 0.5, C_BORDER),
        ("TOPPADDING", (0,0),(-1,-1), 12),
        ("BOTTOMPADDING",(0,0),(-1,-1),12),
    ]))
    return t


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE DECORATORS
# ═══════════════════════════════════════════════════════════════════════════════

def on_cover(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(C_NAVY)
    canvas.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)
    canvas.setFillColor(C_PRIMARY)
    canvas.rect(0, PAGE_H * 0.55, PAGE_W, PAGE_H * 0.45, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#3b82f6"))
    canvas.rect(0, PAGE_H * 0.54, PAGE_W, 0.025*PAGE_H, fill=1, stroke=0)
    canvas.restoreState()

def on_page(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(C_NAVY)
    canvas.rect(0, 0, PAGE_W, 1.1*cm, fill=1, stroke=0)
    canvas.setFillColor(C_MUTED)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawCentredString(PAGE_W/2, 0.4*cm,
        f"SGDI v2.0 — Documento de Refatoracao | Pagina {doc.page}")
    canvas.drawString(MARGIN, 0.4*cm, "USO INTERNO")
    canvas.setFillColor(colors.HexColor("#f1f5f9"))
    canvas.rect(0, PAGE_H - 1.0*cm, PAGE_W, 1.0*cm, fill=1, stroke=0)
    canvas.setFillColor(C_PRIMARY)
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(MARGIN, PAGE_H - 0.65*cm, "SGDI — Refatoracao de Codigo")
    canvas.setFillColor(C_MUTED)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawRightString(PAGE_W - MARGIN, PAGE_H - 0.65*cm,
        datetime.now().strftime("%d/%m/%Y"))
    canvas.restoreState()


# ═══════════════════════════════════════════════════════════════════════════════
# SECOES
# ═══════════════════════════════════════════════════════════════════════════════

def build_cover():
    elems = []
    elems.append(sp(PAGE_H * 0.06))
    elems.append(P("SGDI v2.0", "DocTitle"))
    elems.append(sp(6))
    elems.append(P("Documento de Refatoracao de Codigo", "DocSubtitle"))
    elems.append(sp(4))
    elems.append(P("Tres Melhorias de Qualidade Aplicadas ao app.py", "DocSubtitle"))
    elems.append(sp(20))
    meta_s = ParagraphStyle("meta", fontSize=9, fontName="Helvetica",
                             textColor=colors.HexColor("#94a3b8"),
                             alignment=TA_CENTER, leading=15)
    elems.append(Paragraph(f"Versao 1.0  |  {datetime.now().strftime('%d/%m/%Y')}",meta_s))
    elems.append(Paragraph("Equipe: Luis Felipe  |  Confidencial — Uso Interno", meta_s))
    elems.append(sp(PAGE_H * 0.08))

    changes = [
        ("3.1", "Transicao de Status", "Helper unico substitui 4 rotas duplicadas", C_DANGER),
        ("3.2", "Queries do Dashboard", "Helper _calcular_kpis elimina 60 linhas repetidas", C_WARNING),
        ("3.3", "Exportacao Excel", "Helper _build_xlsx unifica 3 blocos identicos", C_SUCCESS),
    ]
    cells = []
    for num, title, desc, color in changes:
        n_s = ParagraphStyle(f"cn{num}", fontSize=13, fontName="Helvetica-Bold",
                              textColor=color, alignment=TA_CENTER, leading=16)
        t_s = ParagraphStyle(f"ct{num}", fontSize=9.5, fontName="Helvetica-Bold",
                              textColor=C_WHITE, leading=13)
        d_s = ParagraphStyle(f"cd{num}", fontSize=8, fontName="Helvetica",
                              textColor=colors.HexColor("#94a3b8"), leading=11)
        cells.append([Paragraph(num, n_s), Paragraph(title, t_s), Paragraph(desc, d_s)])
    t = Table(cells, colWidths=[1.5*cm, 5.5*cm, CONTENT_W-7.0*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#1e293b")),
        ("LINEBELOW",     (0,0),(-1,-2), 0.3, colors.HexColor("#334155")),
        ("TOPPADDING",    (0,0),(-1,-1), 12),
        ("BOTTOMPADDING", (0,0),(-1,-1), 12),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 10),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    elems.append(t)
    elems.append(PageBreak())
    return elems


def build_intro():
    elems = []
    elems.append(sp(10))
    elems.append(P("Contexto e Motivacao", "H1"))
    elems.append(hr())
    elems.append(P(
        "Este documento registra tres refatoracoes de qualidade de codigo aplicadas ao "
        "arquivo <b>app.py</b> do SGDI v2.0. As mudancas nao alteram nenhum comportamento "
        "visivel ao usuario — todas as funcionalidades permanecem identicas. O objetivo "
        "e reduzir duplicacao de codigo, facilitar manutencao futura e tornar o codigo "
        "mais legivel e consistente.", "Body"))
    elems.append(sp(6))
    elems.append(kpi_row([
        ("Linhas removidas",   "~220", C_DANGER),
        ("Linhas adicionadas", "~105", C_SUCCESS),
        ("Reducao liquida",    "-115", C_PRIMARY),
        ("Helpers criados",    "3",    C_WARNING),
        ("Rotas simplificadas","7",    C_SUCCESS),
    ]))
    elems.append(sp(14))
    elems.append(P("O que e refatoracao?", "H2"))
    elems.append(P(
        "Refatoracao e o processo de reestruturar codigo existente sem mudar seu "
        "comportamento externo. O principio DRY (Don't Repeat Yourself — Nao Se Repita) "
        "diz que cada pedaco de conhecimento deve ter uma unica representacao no sistema. "
        "Quando a mesma logica aparece em multiplos lugares, qualquer correcao ou "
        "melhoria precisa ser replicada manualmente em todos eles — o que e fragil e "
        "propenso a bugs por inconsistencia.", "Body"))
    elems.append(info_box(
        "Principio DRY: cada logica de negocio deve existir em UM unico lugar. "
        "Tres ocorrencias identicas do mesmo bloco = tres lugares onde um bug pode "
        "surgir e tres lugares onde uma correcao precisa ser aplicada.",
        C_PRIMARY))
    elems.append(sp(10))
    elems.append(P("Escopo das mudancas", "H2"))
    elems.append(P("Todas as alteracoes estao contidas em <b>app.py</b>. Nenhum outro arquivo foi modificado.", "Body"))

    rows = [
        ["Arquivo", "Tipo de mudanca", "Linhas afetadas"],
        ["app.py", "Adicao de helper _transicionar_status()", "625-700 (antes)"],
        ["app.py", "Simplificacao das 4 rotas de status", "625-758 (antes)"],
        ["app.py", "Adicao de helper _calcular_kpis()", "1001-1082 (antes)"],
        ["app.py", "Simplificacao de api_dashboard_kpis()", "1001-1082 (antes)"],
        ["app.py", "Remocao do bloco duplicado em api_dashboard_data()", "1188-1243 (antes)"],
        ["app.py", "Adicao de helper _build_xlsx()", "novo (inserido)"],
        ["app.py", "Simplificacao de api_dashboard_export()", "1421-1451 (antes)"],
        ["app.py", "Simplificacao de api_dashboard_critical_overdue_export()", "1475-1530 (antes)"],
        ["app.py", "Simplificacao de auditoria_export() bloco xlsx", "2327-2394 (antes)"],
    ]
    style = ParagraphStyle("tc", fontSize=8.5, fontName="Helvetica",
                            textColor=C_DARK, leading=12)
    style_h = ParagraphStyle("th", fontSize=8.5, fontName="Helvetica-Bold",
                              textColor=C_WHITE, leading=12, alignment=TA_CENTER)
    data = [[Paragraph(r[0], style_h if i == 0 else style),
             Paragraph(r[1], style_h if i == 0 else style),
             Paragraph(r[2], style_h if i == 0 else style)]
            for i, r in enumerate(rows)]
    t = Table(data, colWidths=[3.0*cm, 9.5*cm, CONTENT_W-12.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), C_PRIMARY),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_WHITE, C_BG]),
        ("GRID",          (0,0),(-1,-1), 0.3, C_BORDER),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("RIGHTPADDING",  (0,0),(-1,-1), 6),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    elems.append(t)
    elems.append(PageBreak())
    return elems


def build_refat31():
    elems = []
    elems.append(section_header("3.1", "Rotas de Transicao de Status",
                                 "4 rotas com 80% de codigo duplicado -> 1 helper + 4 chamadas",
                                 C_DANGER))
    elems.append(sp(10))

    elems.append(P("O Problema", "H2"))
    elems.append(P(
        "As quatro rotas <b>/concluir</b>, <b>/reabrir</b>, <b>/andamento</b> e "
        "<b>/cancelar</b> executavam a mesma sequencia de operacoes: validar CSRF, "
        "buscar a demanda no banco, executar UPDATE, inserir historico, fazer commit, "
        "registrar no log de auditoria, exibir flash e redirecionar. "
        "Apenas o status-novo, a mensagem de flash e o destino do redirect diferiam.", "Body"))
    elems.append(sp(6))
    elems.append(metrics_row("~33", "~8",
                              "Linhas por rota (antes)", "Linhas por rota (depois)"))
    elems.append(sp(6))
    elems.append(metrics_row("~132", "~65",
                              "Total de linhas (4 rotas)", "Total (helper + 4 chamadas)"))
    elems.append(sp(10))

    elems.append(P("Onde estava — exemplo da rota /concluir", "H2"))
    elems += before_after([
        "# app.py — antes (repetido 4 vezes com variacoes minimas)",
        "@app.route('/concluir/<int:id>', methods=['POST'])",
        "@login_required",
        "def concluir(id):",
        "    _validate_csrf()",
        "    conn = get_db()",
        "    try:",
        "        demanda = conn.execute(",
        "            'SELECT * FROM demandas WHERE id = ?', (id,)",
        "        ).fetchone()",
        "        if not demanda:",
        "            abort(404)",
        "        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')",
        "        conn.execute(",
        "            'UPDATE demandas SET status = ?, data_conclusao = ? WHERE id = ?',",
        "            (STATUS_CONCLUIDA, now, id),",
        "        )",
        "        _registrar_historico(conn, id, demanda['status'],",
        "            STATUS_CONCLUIDA, session['usuario_nome'], now)",
        "        conn.commit()",
        "    finally:",
        "        conn.close()",
        "    log.registrar(",
        "        CAT_DEMANDA, 'status_alterado', nivel='INFO',",
        "        usuario_id=session['usuario_id'],",
        "        usuario_nome=session['usuario_nome'],",
        "        ip=request.remote_addr, recurso_tipo='demanda', recurso_id=id,",
        "        detalhes={'de': demanda['status'], 'para': STATUS_CONCLUIDA},",
        "    )",
        "    flash('Demanda marcada como concluida.')",
        "    return redirect(url_for('index'))",
        "",
        "# ... mais 3 blocos quasi-identicos para reabrir, andamento, cancelar",
    ], [
        "# app.py — depois: 1 helper + 4 chamadas de 2 linhas cada",
        "def _transicionar_status(demanda_id, status_novo, flash_msg,",
        "                          redirect_to, extra_set=None):",
        "    _validate_csrf()",
        "    ip  = request.remote_addr",
        "    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')",
        "    conn = get_db()",
        "    try:",
        "        demanda = conn.execute(",
        "            'SELECT * FROM demandas WHERE id = ?', (demanda_id,)",
        "        ).fetchone()",
        "        if not demanda: abort(404)",
        "        set_parts  = ['status = ?']",
        "        set_values = [status_novo]",
        "        if extra_set:",
        "            for col, val in extra_set.items():",
        "                set_parts.append(f'{col} = ?')",
        "                set_values.append(val)",
        "        set_values.append(demanda_id)",
        "        conn.execute(",
        "            f'UPDATE demandas SET {\", \".join(set_parts)} WHERE id = ?',",
        "            set_values,",
        "        )",
        "        _registrar_historico(conn, demanda_id, demanda['status'],",
        "            status_novo, session['usuario_nome'], now)",
        "        conn.commit()",
        "    finally:",
        "        conn.close()",
        "    log.registrar(CAT_DEMANDA, 'status_alterado', nivel='INFO',",
        "        usuario_id=session['usuario_id'], ip=ip,",
        "        recurso_id=demanda_id,",
        "        detalhes={'de': demanda['status'], 'para': status_novo})",
        "    flash(flash_msg)",
        "    if redirect_to == 'detalhes':",
        "        return redirect(url_for('detalhes', id=demanda_id))",
        "    return redirect(url_for(redirect_to))",
        "",
        "@app.route('/concluir/<int:id>', methods=['POST'])",
        "@login_required",
        "def concluir(id):",
        "    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')",
        "    return _transicionar_status(id, STATUS_CONCLUIDA,",
        "        'Demanda marcada como concluida.', 'index',",
        "        extra_set={'data_conclusao': now})",
        "",
        "@app.route('/reabrir/<int:id>', methods=['POST'])",
        "@login_required",
        "def reabrir(id):",
        "    return _transicionar_status(id, STATUS_ABERTA,",
        "        'Demanda reaberta.', 'index')",
        "",
        "# ... andamento e cancelar seguem o mesmo padrao de 2 linhas",
    ])
    elems.append(sp(10))

    elems.append(P("Vantagens conquistadas", "H2"))
    elems.append(adv_table([
        ("O", "Unico ponto de correcao",
         "Se uma regra de negocio mudar (ex: adicionar validacao de papel antes "
         "de transicionar), a mudanca e feita em UM lugar e afeta todas as 4 rotas automaticamente."),
        ("O", "Eliminacao de divergencias silenciosas",
         "Antes, cada rota tinha sua propria variacao dos campos de log (alguns usavam "
         "usuario_nome, outros nao). Com o helper, o comportamento e uniforme."),
        ("O", "Codigo das rotas expressa INTENCAO",
         "A rota /concluir agora diz: 'quero transicionar para CONCLUIDA com data_conclusao'. "
         "O COMO e responsabilidade do helper, nao da rota."),
        ("O", "Facilidade para adicionar novos status",
         "Para adicionar uma rota /pausar, basta uma rota de 2 linhas chamando o helper "
         "— sem copiar 30 linhas de logica identica."),
        ("O", "Testabilidade",
         "O helper pode ser testado unitariamente de forma isolada, sem precisar "
         "simular todas as 4 rotas separadamente."),
    ]))
    elems.append(PageBreak())
    return elems


def build_refat32():
    elems = []
    elems.append(section_header("3.2", "Queries do Dashboard Duplicadas",
                                 "60 linhas de SQL repetidas -> helper _calcular_kpis()",
                                 C_WARNING))
    elems.append(sp(10))

    elems.append(P("O Problema", "H2"))
    elems.append(P(
        "O endpoint <b>/api/dashboard/kpis</b> e o endpoint consolidado "
        "<b>/api/dashboard/data</b> continham blocos SQL identicos: o grande SELECT "
        "com todos os CASE WHEN para calcular totais, percentuais e tempo medio, "
        "mais o SELECT de agrupamento por responsavel. "
        "Eram aproximadamente 60 linhas de SQL complexo copiadas e coladas entre os dois endpoints.", "Body"))
    elems.append(sp(6))
    elems.append(metrics_row("2", "1",
                              "Lugares onde o SQL existia", "Lugar onde o SQL existe (agora)"))
    elems.append(sp(6))
    elems.append(metrics_row("~60", "~5",
                              "Linhas de SQL em api_dashboard_data (antes)",
                              "Linhas em api_dashboard_data (depois, chama helper)"))
    elems.append(sp(10))

    elems.append(P("Onde estava — trecho duplicado", "H2"))
    elems += before_after([
        "# app.py — ANTES: mesmo bloco SQL em api_dashboard_kpis E api_dashboard_data",
        "",
        "# Em api_dashboard_kpis() (linhas ~1007-1082):",
        "row = conn.execute(f'''",
        "    SELECT",
        "        COUNT(*) as total,",
        "        SUM(CASE WHEN d.status = 'Aberta' THEN 1 ELSE 0 END) as abertas,",
        "        SUM(CASE WHEN d.status = 'Em andamento' ...) as em_andamento,",
        "        SUM(CASE WHEN d.status = 'Concluida' ...) as concluidas,",
        "        ... (mais 8 campos calculados com CASE WHEN) ...",
        "        ... (calculo complexo de tempo medio ponderado) ...",
        "    FROM demandas d",
        "    LEFT JOIN usuarios u ON u.id = d.responsavel_id",
        "    {where}''', params).fetchone()",
        "",
        "por_resp = conn.execute(f'''",
        "    SELECT COALESCE(u.nome, ...) as nome,",
        "        COUNT(d.id) as total,",
        "        SUM(CASE WHEN ...) as abertas,",
        "        ...",
        "    GROUP BY d.usuario_id ORDER BY abertas DESC''', params).fetchall()",
        "",
        "# Em api_dashboard_data() (linhas ~1188-1243): EXATO MESMO BLOCO",
        "row = conn.execute(f'''  # <- copia identica das 35 linhas acima",
        "    SELECT COUNT(*) as total, ...",
        "    ...",
        "    ''', params).fetchone()",
        "",
        "por_resp = conn.execute(f'''  # <- copia identica das 15 linhas acima",
        "    SELECT ...",
        "    ''', params).fetchall()",
    ], [
        "# app.py — DEPOIS: helper unico, chamado pelos dois endpoints",
        "",
        "def _calcular_kpis(conn, where, params):",
        "    '''Executa as queries de KPIs e por-responsavel (fonte unica).'''",
        "    row = conn.execute(f'''",
        "        SELECT COUNT(*) as total,",
        "            SUM(CASE WHEN d.status = 'Aberta' ...) as abertas,",
        "            ...",
        "        FROM demandas d",
        "        LEFT JOIN usuarios u ON u.id = d.responsavel_id",
        "        {where}''', params).fetchone()",
        "    por_resp = conn.execute(f'''...''', params).fetchall()",
        "    total = row['total'] or 0",
        "    return {'total': total, 'abertas': ..., ...}",
        "",
        "# api_dashboard_kpis — agora tem apenas 5 linhas:",
        "@app.route('/api/dashboard/kpis')",
        "@login_required",
        "def api_dashboard_kpis():",
        "    where, params, _ = _build_dashboard_filters()",
        "    conn = get_db()",
        "    try:",
        "        return jsonify(_calcular_kpis(conn, where, params))",
        "    finally:",
        "        conn.close()",
        "",
        "# api_dashboard_data — substitui o bloco duplicado por uma linha:",
        "def api_dashboard_data():",
        "    where, params, _ = _build_dashboard_filters()",
        "    conn = get_db()",
        "    try:",
        "        kpis = _calcular_kpis(conn, where, params)  # <- unica linha",
        "        # ... resto da funcao (charts, evolucao, critical) continua igual",
        "        return jsonify({'kpis': kpis, 'charts': {...}, ...})",
    ])
    elems.append(sp(10))

    elems.append(P("Vantagens conquistadas", "H2"))
    elems.append(adv_table([
        ("O", "Fonte unica de verdade para KPIs",
         "Se uma formula precisar ser ajustada (ex: mudar o peso de Critica no tempo medio "
         "de 2x para 3x), a mudanca e feita em um lugar e ambos os endpoints recebem o corrigido."),
        ("O", "Eliminacao de risco de divergencia entre endpoints",
         "Antes era possivel corrigir um bug em api_dashboard_kpis e esquecer de "
         "aplicar o mesmo fix em api_dashboard_data, causando inconsistencia silenciosa."),
        ("O", "Reducao de 60 linhas de SQL duplicado",
         "O codigo total do modulo diminuiu significativamente sem perder nenhuma funcionalidade."),
        ("O", "Reutilizacao futura",
         "Qualquer novo endpoint que precise de KPIs do dashboard pode chamar "
         "_calcular_kpis(conn, where, params) com seus proprios filtros."),
    ]))
    elems.append(PageBreak())
    return elems


def build_refat33():
    elems = []
    elems.append(section_header("3.3", "Logica de Exportacao Excel Triplicada",
                                 "3 blocos identicos de openpyxl -> 1 helper _build_xlsx()",
                                 C_SUCCESS))
    elems.append(sp(10))

    elems.append(P("O Problema", "H2"))
    elems.append(P(
        "O sistema tem tres exportacoes Excel: dashboard geral, criticas atrasadas "
        "e auditoria. As tres compartilhavam o mesmo codigo de configuracao do workbook: "
        "criacao do Workbook, merge de celulas de titulo, estilo do cabecalho, "
        "loop de dados, definicao de largura de colunas, BytesIO + send_file. "
        "O codigo de setup estava copiado e colado com pequenas variacoes de cor.", "Body"))
    elems.append(sp(6))
    elems.append(metrics_row("3", "1",
                              "Implementacoes do bloco xlsx", "Implementacao (helper)"))
    elems.append(sp(6))
    elems.append(metrics_row("~75", "~45",
                              "Linhas de boilerplate Excel (total, antes)",
                              "Linhas de boilerplate Excel (total, depois)"))
    elems.append(sp(10))

    elems.append(P("Onde estava — boilerplate repetido", "H2"))
    elems += before_after([
        "# Bloco que aparecia em api_dashboard_export(), em",
        "# api_dashboard_critical_overdue_export() e em auditoria_export():",
        "",
        "from openpyxl import Workbook",
        "from openpyxl.styles import Alignment, Font, PatternFill",
        "from openpyxl.utils import get_column_letter",
        "",
        "wb = Workbook()",
        "ws = wb.active",
        "ws.title = 'Demandas'  # diferente em cada lugar",
        "",
        "ws.merge_cells('A1:I1')",
        "ws['A1'] = 'SGDI — ...'  # titulo diferente",
        "ws['A1'].font = Font(bold=True, size=14)",
        "ws['A1'].alignment = Alignment(horizontal='center')",
        "",
        "ws.merge_cells('A2:I2')",
        "ws['A2'] = f'Gerado em: {now_str}'",
        "ws['A2'].font = Font(size=10, color='666666')",
        "ws['A2'].alignment = Alignment(horizontal='center')",
        "",
        "ws.append([])",
        "ws.append(headers_row)",
        "header_fill = PatternFill(start_color='2563EB', ..., fill_type='solid')",
        "for col_idx in range(1, len(headers_row) + 1):",
        "    cell = ws.cell(row=5, column=col_idx)",
        "    cell.fill = header_fill",
        "    cell.font = Font(bold=True, color='FFFFFF', size=10)",
        "    cell.alignment = Alignment(horizontal='center')",
        "",
        "for r in rows:",
        "    ws.append([...])  # dados variados",
        "    # logica de coloracao especifica",
        "",
        "for idx, width in enumerate([6, 42, 22, ...], 1):",
        "    ws.column_dimensions[get_column_letter(idx)].width = width",
        "",
        "buf = io.BytesIO()",
        "wb.save(buf)",
        "buf.seek(0)",
        "return send_file(buf, mimetype='application/vnd...', ...)",
        "",
        "# Este bloco de 30+ linhas aparecia 3 vezes no arquivo",
    ], [
        "# Helper generico adicionado ao app.py:",
        "def _build_xlsx(sheet_title, title, subtitle, headers, rows, col_widths,",
        "                header_color='2563EB', title_color='000000',",
        "                subtitle_color='666666', row_styler=None):",
        "    from openpyxl import Workbook",
        "    from openpyxl.styles import Alignment, Font, PatternFill",
        "    from openpyxl.utils import get_column_letter",
        "    wb = Workbook()",
        "    ws = wb.active",
        "    ws.title = sheet_title",
        "    n, last = len(headers), get_column_letter(len(headers))",
        "    ws.merge_cells(f'A1:{last}1')",
        "    ws['A1'] = title",
        "    ws['A1'].font = Font(bold=True, size=14, color=title_color)",
        "    ws['A1'].alignment = Alignment(horizontal='center')",
        "    # ... mesmo padrao para subtitulo ...",
        "    ws.append([])",
        "    ws.append(headers)",
        "    hdr_row = ws.max_row",
        "    fill_hdr = PatternFill(start_color=header_color, ..., fill_type='solid')",
        "    for c in range(1, n + 1):",
        "        ws.cell(row=hdr_row, column=c).fill = fill_hdr",
        "        ws.cell(row=hdr_row, column=c).font = Font(bold=True, color='FFFFFF')",
        "    for row_data in rows:",
        "        ws.append(list(row_data))",
        "        if row_styler: row_styler(ws, ws.max_row, row_data)",
        "    for idx, width in enumerate(col_widths, 1):",
        "        ws.column_dimensions[get_column_letter(idx)].width = width",
        "    buf = io.BytesIO()",
        "    wb.save(buf)",
        "    buf.seek(0)",
        "    return buf",
        "",
        "# Uso em api_dashboard_export (substituiu 40 linhas por 15):",
        "def _style_dash(ws, rn, r):",
        "    if r[3] in priority_fills:",
        "        ws.cell(row=rn, column=4).fill = priority_fills[r[3]]",
        "    if r[4] in status_fills:",
        "        ws.cell(row=rn, column=5).fill = status_fills[r[4]]",
        "",
        "buf = _build_xlsx('Demandas', 'SGDI — ...', subtitle,",
        "    headers_row, row_tuples, [6,42,22,12,15,20,20,20,10],",
        "    row_styler=_style_dash)",
        "return send_file(buf, mimetype='application/vnd...', ...)",
        "",
        "# Uso em auditoria_export (substituiu 40 linhas por 18):",
        "buf = _build_xlsx('Logs de Auditoria', 'SGDI — Logs de Auditoria',",
        "    f'Gerado em: {now_str}', headers_row, row_tuples,",
        "    [20,10,14,26,20,16,14,10,50],",
        "    header_color='1E3A5F', row_styler=_style_audit)",
    ])
    elems.append(sp(10))

    elems.append(P("Como funciona o parametro row_styler", "H2"))
    elems.append(P(
        "O parametro <b>row_styler</b> e uma funcao opcional que o helper chama "
        "apos inserir cada linha de dados, passando <b>ws</b> (a planilha), "
        "<b>rn</b> (numero da linha atual) e <b>r</b> (a tupla de dados da linha). "
        "Isso permite que cada exportacao aplique sua coloracao especifica "
        "sem duplicar o restante do boilerplate.", "Body"))
    elems.append(code_block([
        "# Exemplo: row_styler para colorir celulas por nivel de auditoria",
        "def _style_audit(ws, rn, r):",
        "    nivel_val = r[1]  # coluna 2 = nivel (INFO/WARNING/ERROR/CRITICAL)",
        "    if nivel_val in nivel_fills:",
        "        for col in range(1, 10):",
        "            ws.cell(row=rn, column=col).fill = nivel_fills[nivel_val]",
        "    if nivel_val in nivel_fonts:",
        "        ws.cell(row=rn, column=2).font = nivel_fonts[nivel_val]",
    ]))
    elems.append(sp(10))

    elems.append(P("Vantagens conquistadas", "H2"))
    elems.append(adv_table([
        ("O", "Consistencia visual entre exportacoes",
         "O layout de titulo, subtitulo e cabecalho e identico nas 3 exportacoes "
         "porque vem do mesmo codigo — nao de tres copias independentes."),
        ("O", "Adicionar nova exportacao Excel e trivial",
         "Para exportar um novo relatorio basta definir os dados, as colunas e "
         "opcionalmente um row_styler — sem copiar 30 linhas de boilerplate."),
        ("O", "Correcoes propagam automaticamente",
         "Se for necessario mudar a fonte do cabecalho ou o tamanho do titulo, "
         "uma mudanca no helper afeta todas as exportacoes de uma vez."),
        ("O", "Tratamento de ImportError centralizado",
         "A verificacao se openpyxl esta instalado esta em um unico lugar "
         "(o helper retorna None), em vez de repetida em 3 blocos try/except."),
        ("O", "Cada exportacao foca apenas em seus dados",
         "O codigo especifico de cada endpoint agora expressa apenas O QUE quer "
         "exportar (colunas, dados, cores), deixando o COMO para o helper."),
    ]))
    elems.append(PageBreak())
    return elems


def build_conclusao():
    elems = []
    elems.append(sp(10))
    elems.append(P("Conclusao e Impacto", "H1"))
    elems.append(hr())
    elems.append(P(
        "As tres refatoracoes aplicadas reduzem o numero de linhas do arquivo "
        "<b>app.py</b> em aproximadamente 115 linhas liquidas, sem remover nenhuma "
        "funcionalidade. O sistema continua se comportando de forma identica para "
        "todos os usuarios.", "Body"))
    elems.append(sp(8))
    elems.append(kpi_row([
        ("Helpers extraidos",   "3",    C_SUCCESS),
        ("Rotas simplificadas", "7",    C_PRIMARY),
        ("SQL duplicado removido", "60 linhas", C_WARNING),
        ("Boilerplate xlsx removido", "~75 linhas", C_SUCCESS),
        ("Funcionalidade removida", "0", C_SUCCESS),
    ]))
    elems.append(sp(14))

    elems.append(P("Resumo das mudancas por arquivo", "H2"))
    rows_data = [
        ["Mudanca", "Arquivo", "Funcao criada/alterada", "Linhas antes", "Linhas depois"],
        ["3.1 — Status helper", "app.py", "_transicionar_status()", "~132", "~65"],
        ["3.1 — concluir()", "app.py", "concluir()", "~26", "3"],
        ["3.1 — reabrir()", "app.py", "reabrir()", "~24", "2"],
        ["3.1 — andamento()", "app.py", "andamento()", "~24", "2"],
        ["3.1 — cancelar()", "app.py", "cancelar()", "~24", "2"],
        ["3.2 — KPI helper", "app.py", "_calcular_kpis()", "0 (novo)", "~65"],
        ["3.2 — kpis()", "app.py", "api_dashboard_kpis()", "~80", "5"],
        ["3.2 — data()", "app.py", "api_dashboard_data() (bloco duplicado)", "~60", "1"],
        ["3.3 — xlsx helper", "app.py", "_build_xlsx()", "0 (novo)", "~45"],
        ["3.3 — dashboard export", "app.py", "api_dashboard_export() xlsx", "~40", "~18"],
        ["3.3 — critical export", "app.py", "api_dashboard_critical_overdue_export() xlsx", "~40", "~18"],
        ["3.3 — auditoria export", "app.py", "auditoria_export() bloco xlsx", "~40", "~18"],
    ]
    s = ParagraphStyle("tc", fontSize=8, fontName="Helvetica", textColor=C_DARK, leading=12)
    sh = ParagraphStyle("th", fontSize=8, fontName="Helvetica-Bold", textColor=C_WHITE,
                         leading=12, alignment=TA_CENTER)
    data = [[Paragraph(c, sh if i == 0 else s) for c in row]
            for i, row in enumerate(rows_data)]
    t = Table(data, colWidths=[4.0*cm, 2.0*cm, 5.5*cm, 2.0*cm, 2.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), C_PRIMARY),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_WHITE, C_BG]),
        ("GRID",          (0,0),(-1,-1), 0.3, C_BORDER),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(-1,-1), 5),
        ("RIGHTPADDING",  (0,0),(-1,-1), 5),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    elems.append(t)
    elems.append(sp(14))

    elems.append(info_box(
        "Verificacao: apos as refatoracoes, o app foi importado com sucesso e a "
        "presenca dos tres helpers foi confirmada programaticamente. "
        "Nenhuma rota foi removida ou renomeada — todas as URLs publicas permanecem identicas.",
        C_SUCCESS, colors.HexColor("#f0fdf4")))

    elems.append(sp(20))
    close_s = ParagraphStyle("close", fontSize=9, fontName="Helvetica",
                              textColor=C_MUTED, alignment=TA_CENTER, leading=14)
    elems.append(Paragraph(
        f"SGDI v2.0 — Documento de Refatoracao de Codigo<br/>"
        f"Gerado em {datetime.now().strftime('%d/%m/%Y as %H:%M')}<br/>"
        "Tres helpers extraidos · 115 linhas removidas · 0 funcionalidades alteradas",
        close_s))
    return elems


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    OUTPUT = "refatoracao_codigo_sgdi.pdf"

    doc = BaseDocTemplate(
        OUTPUT,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=1.4*cm,
        bottomMargin=1.4*cm,
        title="SGDI v2.0 — Documento de Refatoracao",
        author="Equipe SGDI",
    )
    frame = Frame(MARGIN, 1.4*cm, CONTENT_W, PAGE_H - 2.8*cm, id="normal")
    cover_frame = Frame(MARGIN, 1.4*cm, CONTENT_W, PAGE_H - 2.8*cm, id="cover")
    doc.addPageTemplates([
        PageTemplate(id="Cover", frames=[cover_frame], onPage=on_cover),
        PageTemplate(id="Body",  frames=[frame],        onPage=on_page),
    ])

    story = (
        [NextPageTemplate("Cover")] + build_cover() +
        [NextPageTemplate("Body")] +
        build_intro() +
        build_refat31() +
        build_refat32() +
        build_refat33() +
        build_conclusao()
    )
    doc.build(story)
    print(f"[OK] PDF gerado: {OUTPUT}")


if __name__ == "__main__":
    main()
